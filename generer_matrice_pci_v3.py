# =============================================================================
#  GÉNÉRATEUR v3 — Matrice de Suivi Remédiation PCI
#  Amundi Immobilier — Opérations & Conformité Externalisations
#  Usage  : python generer_matrice_pci_v3.py
#  Sortie : Matrice_PCI_Remediation_v3.xlsx
#  Prérequis : pip install openpyxl
#  Onglets : MODE D'EMPLOI | DATABASE | REMÉDIATION | CONTRÔLES LOD1 | SYNTHÈSE
# =============================================================================

import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

# ── Palette Amundi ──────────────────────────────────────────────────────────
NAVY      = "001C4B"; BLUE      = "009EE0"; LIGHT_BLUE = "CCE5F3"
MID_BLUE  = "66B3DB"; WHITE     = "FFFFFF"; GREY_BG    = "F5F7FA"
GREEN_LT  = "D6F0DD"; ORANGE_LT = "FFE5B4"; RED_LT     = "FFD7D7"
YELLOW_LT = "FFFDE7"; GREY_BDR  = "BFCAD4"; NA_BG      = "EEEEEE"

# ── Helpers ─────────────────────────────────────────────────────────────────
def hdr(c, t, bg=NAVY, fg=WHITE, sz=10, bold=True):
    c.value = t
    c.font = Font(name="Arial", bold=bold, color=fg, size=sz)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def sub(c, t, bg=LIGHT_BLUE, fg=NAVY, sz=9):
    c.value = t
    c.font = Font(name="Arial", bold=True, color=fg, size=sz)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def std(c, v="", sz=9, bold=False, bg=WHITE, center=False, italic=False):
    c.value = v
    c.font = Font(name="Arial", size=sz, bold=bold, italic=italic)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center" if center else "left",
                             vertical="center", wrap_text=True)

def brd(ws, r1, r2, c1, c2):
    s = Side(style="thin", color=GREY_BDR)
    b = Border(left=s, right=s, top=s, bottom=s)
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
        for c in row:
            c.border = b

def dv(ws, formula, sqref):
    v = DataValidation(type="list", formula1=formula,
                       allow_blank=True, showErrorMessage=True)
    v.sqref = sqref
    ws.add_data_validation(v)

def rag(ws, col, r1, r2):
    rng = f"{col}{r1}:{col}{r2}"
    ws.conditional_formatting.add(rng, CellIsRule("equal", ['"Vert"'],
        fill=PatternFill("solid", fgColor=GREEN_LT),
        font=Font(name="Arial", size=9, color="00703C")))
    ws.conditional_formatting.add(rng, CellIsRule("equal", ['"Orange"'],
        fill=PatternFill("solid", fgColor=ORANGE_LT),
        font=Font(name="Arial", size=9, color="7D4500")))
    ws.conditional_formatting.add(rng, CellIsRule("equal", ['"Rouge"'],
        fill=PatternFill("solid", fgColor=RED_LT),
        font=Font(name="Arial", size=9, color="C00000")))

def date_mfc(ws, col, r1, r2):
    rng = f"{col}{r1}:{col}{r2}"
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'AND(ISNUMBER({col}{r1}),{col}{r1}<TODAY())'],
        fill=PatternFill("solid", fgColor=RED_LT)))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'AND(ISNUMBER({col}{r1}),{col}{r1}>=TODAY(),{col}{r1}<=TODAY()+30)'],
        fill=PatternFill("solid", fgColor=ORANGE_LT)))

def fill_row(ws, r, n_cols, bg):
    for j in range(1, n_cols + 1):
        c = ws.cell(row=r, column=j)
        if c.fill.fgColor.rgb == "00000000":
            c.fill = PatternFill("solid", fgColor=bg)

TODAY_STR = datetime.date.today().strftime("%d/%m/%Y")
wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# DATABASE
# ══════════════════════════════════════════════════════════════════════════════
ws_db = wb.active
ws_db.title = "DATABASE"
ws_db.sheet_view.showGridLines = False

ws_db.merge_cells("A1:I1")
hdr(ws_db["A1"],
    "DATABASE — Référentiel Prestataires  |  Toute modification ici se répercute automatiquement",
    sz=11)
ws_db.row_dimensions[1].height = 28

db_cols = [
    ("A", "ID Easy",                  16),
    ("B", "Prestataire",              28),
    ("C", "Criticité\n(PCI/PS/PCMG)", 14),
    ("D", "Intragroupe\n(Oui/Non)",   13),
    ("E", "Date fin\ncontrat",         14),
    ("F", "Contact\nMétier",           22),
    ("G", "Mail\nMétier",              30),
    ("H", "Contact\nPrestataire",      22),
    ("I", "Mail\nPrestataire",         30),
]
for col, title, width in db_cols:
    sub(ws_db[f"{col}2"], title)
    ws_db.column_dimensions[col].width = width
ws_db.row_dimensions[2].height = 34

prestataires = [
    ("CTR126384", "Ad Astra",              "PCI",  "Non"),
    ("CTR068683", "AERIUM",                "PS",   "Non"),
    ("TBD",       "Alter Domus Lux",       "PCI",  "Non"),
    ("CTR091882", "Amundi Asset Management","PCI", "Oui"),
    ("CTR068687", "BNPP RE",               "PCMG", "Non"),
    ("CTR068688", "CA Immo",               "PCMG", "Oui"),
    ("CTR148741", "CA Immo",               "PCI",  "Oui"),
    ("CTR068691", "CACEIS",                "PCMG", "Oui"),
    ("CTR068694", "CDC SF",                "PCI",  "Non"),
    ("CTR068696", "Constructa",            "PS",   "Non"),
    ("CTR091883", "CPR ASSET MANAGEMENT",  "PCI",  "Oui"),
    ("CTR116798", "Deltager",              "PCI",  "Oui"),
    ("CTR068698", "Dim4",                  "PCI",  "Non"),
    ("CTR123300", "IC",                    "PCI",  "Non"),
    ("CTR068706", "IDIA",                  "PCI",  "Oui"),
    ("CTR068714", "Mercialys",             "PS",   "Non"),
    ("CTR068716", "Nexity",                "PS",   "Oui"),
    ("CTR129527", "PACK SOLUTIONS",        "PCI",  "Non"),
    ("CTR068724", "SCC",                   "PS",   "Non"),
    ("CTR068726", "SGSS",                  "PS",   "Non"),
    ("CTR077606", "Sienna GER",            "PCI",  "Non"),
    ("CTR068702", "Sienna PB",             "PCI",  "Non"),
    ("CTR068729", "Sinteo",                "PCI",  "Non"),
    ("CTR068732", "Sudeco",                "PS",   "Non"),
    ("CTR068720", "Telamon",               "PS",   "Non"),
    ("CTR068737", "Telmma",                "PCI",  "Non"),
    ("CTR068690", "UPTEVIA",               "PCI",  "Oui"),
    ("CTR136718", "Yardi",                 "PCI",  "Non"),
]
N = len(prestataires)

for i, (id_easy, nom, crit, intra) in enumerate(prestataires, start=3):
    r = i
    ws_db.row_dimensions[r].height = 20
    bg = WHITE if r % 2 == 0 else GREY_BG
    for j, val in enumerate([id_easy, nom, crit, intra, "", "", "", "", ""], start=1):
        c = ws_db.cell(row=r, column=j, value=val)
        c.font = Font(name="Arial", size=9)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(
            horizontal="left" if j in [1, 2, 6, 7, 8, 9] else "center",
            vertical="center")
    ws_db[f"E{r}"].number_format = "DD/MM/YYYY"

dv(ws_db, '"PCI,PS,PCMG"', "C3:C500")
dv(ws_db, '"Oui,Non"',      "D3:D500")
brd(ws_db, 1, N + 2, 1, 9)
ws_db.freeze_panes = "A3"
ws_db.sheet_properties.tabColor = NAVY

# ══════════════════════════════════════════════════════════════════════════════
# REMÉDIATION
# Colonnes : Prestataire | Criticité | Intragroupe |
#            DO | QECI | Avis Risques | Stratégie Sortie |
#            EASY complété | Contrat récent | Remédiation EBA |
#            Statut global | Commentaire
# ══════════════════════════════════════════════════════════════════════════════
ws_rem = wb.create_sheet("REMÉDIATION")
ws_rem.sheet_view.showGridLines = False

ws_rem.merge_cells("A1:L1")
hdr(ws_rem["A1"],
    f"REMÉDIATION — Plan de mise en conformité  |  MàJ : {TODAY_STR}",
    sz=11)
ws_rem.row_dimensions[1].height = 28

# Groupe headers
for rng, txt, bg in [
    ("A2:C2", "IDENTIFICATION",           NAVY),
    ("D2:G2", "DOCUMENTS OBLIGATOIRES",   MID_BLUE),
    ("H2:J2", "CONFORMITÉ OPÉRATIONNELLE",MID_BLUE),
    ("K2:L2", "BILAN",                    NAVY),
]:
    ws_rem.merge_cells(rng)
    hdr(ws_rem[rng.split(":")[0]], txt, bg=bg, sz=9)
ws_rem.row_dimensions[2].height = 18

rem_cols = [
    ("A", "Prestataire",         26),
    ("B", "Criticité",           10),
    ("C", "Intragroupe",         10),
    ("D", "DO\nprésent ?",       12),
    ("E", "QECI\nprésent ?",     12),
    ("F", "Avis\nRisques",       12),
    ("G", "Stratégie\nSortie",   12),
    ("H", "EASY\ncomplété ?",    12),
    ("I", "Contrat\nrécent ?",   12),
    ("J", "Remédiation\nEBA",    12),
    ("K", "Statut\nglobal",      11),
    ("L", "Commentaire / Action",36),
]
for col, title, width in rem_cols:
    sub(ws_rem[f"{col}3"], title)
    ws_rem.column_dimensions[col].width = width
ws_rem.row_dimensions[3].height = 34

# Note de lecture
ws_rem.merge_cells("A4:L4")
c = ws_rem["A4"]
c.value = (
    "ℹ  DO = N/A prestataires existants  |  Avis Risques = PCI seulement  |  "
    "Stratégie Sortie = PCI Hors Groupe seulement  |  "
    "Remédiation EBA = Tous prestataires (PCI + PS + PCMG)  |  "
    "Contrat récent = saisie manuelle"
)
c.font = Font(name="Arial", size=8, italic=True, color=NAVY)
c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_rem.row_dimensions[4].height = 20

for i in range(N):
    r  = i + 5
    db = i + 3      # ligne correspondante dans DATABASE (commence à 3)
    bg = WHITE if r % 2 == 0 else GREY_BG
    ws_rem.row_dimensions[r].height = 22

    # Identification — tirée de DATABASE
    ws_rem[f"A{r}"] = f'=IFERROR(DATABASE!B{db},"")'
    ws_rem[f"B{r}"] = f'=IFERROR(DATABASE!C{db},"")'
    ws_rem[f"C{r}"] = f'=IFERROR(DATABASE!D{db},"")'

    # Contrat récent — saisie manuelle

    # Statut global — Rouge si 1 Non, Orange si 1 En cours, Vert sinon
    # (on ignore les cellules vides et les N/A)
    cols_check = [f"{col}{r}" for col in ["D", "E", "F", "G", "H", "I", "J"]]
    check_range = ",".join(cols_check)
    ws_rem[f"K{r}"] = (
        f'=IF(A{r}="","",IF(COUNTIF({",".join(cols_check[:3])}{",".join([""])}'
        f'D{r}:J{r},"Non")>0,"Rouge",'
        f'IF(COUNTIF(D{r}:J{r},"En cours")>0,"Orange","Vert")))'
    )
    # Simplified formula
    ws_rem[f"K{r}"] = (
        f'=IF(A{r}="","",IF(COUNTIF(D{r}:J{r},"Non")>0,"Rouge",'
        f'IF(COUNTIF(D{r}:J{r},"En cours")>0,"Orange","Vert")))'
    )

    # Style toutes cellules
    for col, _, _ in rem_cols:
        c = ws_rem[f"{col}{r}"]
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(
            horizontal="left" if col in ["A", "L"] else "center",
            vertical="center", wrap_text=True)
        c.fill = PatternFill("solid", fgColor=bg)

# Validations listes déroulantes
ONC = '"Oui,Non,N/A"'
ONN = '"Oui,Non,N/A"'
r1, r2 = 5, 5 + N
dv(ws_rem, ONC, f"D{r1}:D{r2}")   # DO
dv(ws_rem, ONC, f"E{r1}:E{r2}")   # QECI
dv(ws_rem, ONN, f"F{r1}:F{r2}")   # Avis Risques (N/A si pas PCI)
dv(ws_rem, ONN, f"G{r1}:G{r2}")   # Stratégie Sortie (N/A si pas PCI HG)
dv(ws_rem, ONC, f"H{r1}:H{r2}")   # EASY complété
dv(ws_rem, ONC, f"I{r1}:I{r2}")   # Contrat récent
dv(ws_rem, ONC, f"J{r1}:J{r2}")   # Remédiation EBA

# MFC statut global + colonne Contrat récent
rag(ws_rem, "K", r1, r2)
# Contrat récent : MFC rouge si "Non"
rng_i = f"I{r1}:I{r2}"
ws_rem.conditional_formatting.add(rng_i, CellIsRule("equal", ['"Non"'],
    fill=PatternFill("solid", fgColor=RED_LT),
    font=Font(name="Arial", size=9, color="C00000")))
ws_rem.conditional_formatting.add(rng_i, CellIsRule("equal", ['"Oui"'],
    fill=PatternFill("solid", fgColor=GREEN_LT),
    font=Font(name="Arial", size=9, color="00703C")))
ws_rem.conditional_formatting.add(rng_i, CellIsRule("equal", ['"N/A"'],
    fill=PatternFill("solid", fgColor=NA_BG),
    font=Font(name="Arial", size=9, color="888888")))

brd(ws_rem, 1, r2, 1, 12)
ws_rem.freeze_panes = "A5"
ws_rem.sheet_properties.tabColor = MID_BLUE

# ══════════════════════════════════════════════════════════════════════════════
# CONTRÔLES LOD1 — L1-08 + L1-09 uniquement
# ══════════════════════════════════════════════════════════════════════════════
ws_ctrl = wb.create_sheet("CONTRÔLES LOD1")
ws_ctrl.sheet_view.showGridLines = False

ws_ctrl.merge_cells("A1:P1")
hdr(ws_ctrl["A1"],
    "CONTRÔLES LOD1 — L1-08 Événements Majeurs + L1-09 Tests PCA",
    sz=11)
ws_ctrl.row_dimensions[1].height = 26

ws_ctrl.merge_cells("A2:P2")
c = ws_ctrl["A2"]
c.value = (
    "▶  Macros VBA :  [📧 LANCER L1-08]  [📧 RELANCES L1-08]  "
    "[📧 LANCER L1-09]  [📧 RELANCES L1-09]  — Créent des brouillons Outlook"
)
c.font = Font(name="Arial", size=9, italic=True, color=NAVY)
c.fill = PatternFill("solid", fgColor=YELLOW_LT)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_ctrl.row_dimensions[2].height = 20

# Groupes
for rng, txt, bg in [
    ("A3:C3", "IDENTIFICATION",                    NAVY),
    ("D3:I3", "L1-08 | Événements Majeurs (Semestriel — PCI HG)", NAVY),
    ("J3:P3", "L1-09 | Tests PCA (Annuel T1 — Toutes PCI)",       MID_BLUE),
]:
    ws_ctrl.merge_cells(rng)
    hdr(ws_ctrl[rng.split(":")[0]], txt, bg=bg, sz=9)
ws_ctrl.row_dimensions[3].height = 20

ctrl_cols = [
    ("A", "Prestataire",               26),
    ("B", "Criticité",                  9),
    ("C", "Intragroupe",                9),
    # L1-08
    ("D", "S1\nReporting reçu ?",      13),
    ("E", "S1\nStatut",                10),
    ("F", "S1\nCommentaire",           24),
    ("G", "S2\nReporting reçu ?",      13),
    ("H", "S2\nStatut",                10),
    ("I", "S2\nCommentaire",           24),
    # L1-09
    ("J", "Bilan PCA\nreçu ?",         13),
    ("K", "Date test\nPCA",            13),
    ("L", "Anomalies ?",               11),
    ("M", "Statut\nL1-09",             10),
    ("N", "⏱ Prochain\nPCA",           13),
    ("O", "Commentaire\nL1-09",        24),
    ("P", "Statut\nglobal ctrl",       12),
]
for col, title, width in ctrl_cols:
    sub(ws_ctrl[f"{col}4"], title)
    ws_ctrl.column_dimensions[col].width = width
ws_ctrl.row_dimensions[4].height = 36

for i in range(N):
    r  = i + 5
    db = i + 3
    bg = WHITE if r % 2 == 0 else GREY_BG
    ws_ctrl.row_dimensions[r].height = 22

    ws_ctrl[f"A{r}"] = f'=IFERROR(DATABASE!B{db},"")'
    ws_ctrl[f"B{r}"] = f'=IFERROR(DATABASE!C{db},"")'
    ws_ctrl[f"C{r}"] = f'=IFERROR(DATABASE!D{db},"")'

    # L1-08 statuts
    ws_ctrl[f"E{r}"] = (
        f'=IF(A{r}="","",IF(D{r}="","—",'
        f'IF(D{r}="Oui","Vert","Rouge")))'
    )
    ws_ctrl[f"H{r}"] = (
        f'=IF(A{r}="","",IF(G{r}="","—",'
        f'IF(G{r}="Oui","Vert","Rouge")))'
    )

    # L1-09 statut + prochain
    ws_ctrl[f"M{r}"] = (
        f'=IF(A{r}="","",IF(J{r}="","—",'
        f'IF(AND(J{r}="Oui",L{r}="Non"),"Vert",'
        f'IF(AND(J{r}="Oui",L{r}="Oui"),"Orange","Rouge"))))'
    )
    ws_ctrl[f"N{r}"] = (
        f'=IF(K{r}="","À planifier",'
        f'DATE(YEAR(K{r})+1,3,31))'
    )
    ws_ctrl[f"N{r}"].number_format = "DD/MM/YYYY"

    # Statut global contrôles
    ws_ctrl[f"P{r}"] = (
        f'=IF(A{r}="","",IF(OR(E{r}="Rouge",H{r}="Rouge",M{r}="Rouge"),"Rouge",'
        f'IF(OR(E{r}="Orange",H{r}="Orange",M{r}="Orange"),"Orange","Vert")))'
    )

    for col, _, _ in ctrl_cols:
        c = ws_ctrl[f"{col}{r}"]
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(
            horizontal="left" if col in ["A", "F", "I", "O"] else "center",
            vertical="center", wrap_text=True)
        c.fill = PatternFill("solid", fgColor=bg)

    ws_ctrl[f"K{r}"].number_format = "DD/MM/YYYY"

r1c, r2c = 5, 5 + N
dv(ws_ctrl, '"Oui,Non,En attente"', f"D{r1c}:D{r2c}")
dv(ws_ctrl, '"Oui,Non,En attente"', f"G{r1c}:G{r2c}")
dv(ws_ctrl, '"Oui,Non,En attente"', f"J{r1c}:J{r2c}")
dv(ws_ctrl, '"Oui,Non"',            f"L{r1c}:L{r2c}")

for col in ["E", "H", "M", "P"]:
    rag(ws_ctrl, col, r1c, r2c)
date_mfc(ws_ctrl, "N", r1c, r2c)

brd(ws_ctrl, 3, r2c, 1, 16)
ws_ctrl.freeze_panes = "A5"
ws_ctrl.sheet_properties.tabColor = NAVY

# ══════════════════════════════════════════════════════════════════════════════
# SYNTHÈSE
# ══════════════════════════════════════════════════════════════════════════════
ws_syn = wb.create_sheet("SYNTHÈSE")
ws_syn.sheet_view.showGridLines = False

ws_syn.merge_cells("A1:J1")
hdr(ws_syn["A1"],
    f"SYNTHÈSE — Tableau de bord remédiation  |  MàJ : {TODAY_STR}",
    sz=12)
ws_syn.row_dimensions[1].height = 28

syn_cols = [
    ("A", "Prestataire",          28),
    ("B", "ID Easy",              14),
    ("C", "Criticité",            10),
    ("D", "Intragroupe",          10),
    ("E", "Statut\nRemédiation",  12),
    ("F", "L1-08\nS1",            10),
    ("G", "L1-08\nS2",            10),
    ("H", "L1-09\nPCA",           10),
    ("I", "⚠ Risque\nglobal",     12),
    ("J", "Commentaire / Action", 36),
]
for col, title, width in syn_cols:
    sub(ws_syn[f"{col}2"], title)
    ws_syn.column_dimensions[col].width = width
ws_syn.row_dimensions[2].height = 30

for i in range(N):
    r    = i + 3
    db   = i + 3
    rem  = i + 5   # ligne dans REMÉDIATION (commence à 5)
    ctrl = i + 5   # ligne dans CONTRÔLES LOD1

    bg = WHITE if r % 2 == 0 else GREY_BG
    ws_syn.row_dimensions[r].height = 20

    ws_syn[f"A{r}"] = f'=IFERROR(DATABASE!B{db},"")'
    ws_syn[f"B{r}"] = f'=IFERROR(DATABASE!A{db},"")'
    ws_syn[f"C{r}"] = f'=IFERROR(DATABASE!C{db},"")'
    ws_syn[f"D{r}"] = f'=IFERROR(DATABASE!D{db},"")'
    ws_syn[f"E{r}"] = f'=IFERROR(REMÉDIATION!K{rem},"—")'
    ws_syn[f"F{r}"] = f"=IFERROR('CONTRÔLES LOD1'!E{ctrl},\"—\")"
    ws_syn[f"G{r}"] = f"=IFERROR('CONTRÔLES LOD1'!H{ctrl},\"—\")"
    ws_syn[f"H{r}"] = f"=IFERROR('CONTRÔLES LOD1'!M{ctrl},\"—\")"
    ws_syn[f"I{r}"] = (
        f'=IF(A{r}="","",IF(OR(E{r}="Rouge",F{r}="Rouge",G{r}="Rouge",H{r}="Rouge"),"Rouge",'
        f'IF(OR(E{r}="Orange",F{r}="Orange",G{r}="Orange",H{r}="Orange"),"Orange","Vert")))'
    )

    for col, _, _ in syn_cols:
        c = ws_syn[f"{col}{r}"]
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(
            horizontal="left" if col == "J" else "center",
            vertical="center")
        c.fill = PatternFill("solid", fgColor=bg)

for col in ["E", "F", "G", "H", "I"]:
    rag(ws_syn, col, 3, 3 + N)
brd(ws_syn, 2, 3 + N, 1, 10)
ws_syn.freeze_panes = "A3"
ws_syn.sheet_properties.tabColor = "FF0000"

# ══════════════════════════════════════════════════════════════════════════════
# MODE D'EMPLOI
# ══════════════════════════════════════════════════════════════════════════════
ws_me = wb.create_sheet("MODE D'EMPLOI")
ws_me.sheet_view.showGridLines = False
ws_me.column_dimensions["A"].width = 3
ws_me.column_dimensions["B"].width = 28
ws_me.column_dimensions["C"].width = 62

ws_me.merge_cells("B1:C1")
hdr(ws_me["B1"],
    f"MODE D'EMPLOI — Matrice Remédiation PCI v3  |  Amundi Immobilier  |  {TODAY_STR}",
    sz=11)
ws_me.row_dimensions[1].height = 28

row = 3
ws_me.merge_cells(f"B{row}:C{row}")
hdr(ws_me[f"B{row}"], "STRUCTURE DU FICHIER", bg=NAVY, sz=10)
ws_me.row_dimensions[row].height = 22
row += 1

onglets = [
    ("DATABASE",       "Référentiel des 28 prestataires. 9 colonnes : ID Easy, Prestataire, "
                       "Criticité (PCI/PS/PCMG), Intragroupe, Date fin contrat, Contact Métier, "
                       "Mail Métier, Contact Prestataire, Mail Prestataire. "
                       "Toute modification ici se répercute automatiquement dans tous les onglets."),
    ("REMÉDIATION",    "Suivi de la mise en conformité. 1 ligne par prestataire. "
                       "Colonnes à cocher : DO · QECI · Avis Risques · Stratégie Sortie · "
                       "EASY complété · Contrat récent (calculé) · Remédiation EBA. "
                       "Statut global RAG calculé automatiquement."),
    ("CONTRÔLES LOD1", "L1-08 (Événements Majeurs, semestriel, PCI Hors Groupe) et "
                       "L1-09 (Tests PCA, annuel T1, toutes PCI). "
                       "Statuts RAG calculés. Macros VBA pour campagnes mail Outlook."),
    ("SYNTHÈSE",       "Tableau de bord consolidé RAG. Lecture seule — mise à jour automatique."),
]
for nom, desc in onglets:
    ws_me.row_dimensions[row].height = 40
    bg = WHITE if row % 2 == 0 else GREY_BG
    c = ws_me[f"B{row}"]
    c.value = nom
    c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    std(ws_me[f"C{row}"], desc, bg=bg, sz=9)
    row += 1

row += 1
for titre, contenu in [
    ("REMÉDIATION — guide de saisie",
     "DO : N/A pour prestataires existants · Obligatoire pour nouveaux contrats\n"
     "Avis Risques : cocher N/A pour PS et PCMG\n"
     "Stratégie Sortie : cocher N/A si Intragroupe = Oui\n"
     "Remédiation EBA : toutes les lignes PCI (Intra et HG)\n"
     "Contrat récent : calculé automatiquement depuis la date fin contrat (DATABASE col E)"),
    ("CONTRÔLES LOD1 — procédure L1-08 (semestriel)",
     "Périmètre : PCI Hors Groupe  |  Deadline : 30 juin (S1) · 30 décembre (S2)\n"
     "1. Cliquer [LANCER L1-08] → brouillons Outlook créés pour chaque responsable métier\n"
     "2. Vérifier et envoyer chaque mail\n"
     "3. À réception : saisir Oui/Non dans col D (S1) ou G (S2)\n"
     "4. À J-7, J-3, J-1 : cliquer [RELANCES L1-08] → relances auto pour non-répondants"),
    ("CONTRÔLES LOD1 — procédure L1-09 (annuel T1)",
     "Périmètre : Toutes PCI  |  Deadline : 31 mars\n"
     "1. Cliquer [LANCER L1-09] → brouillons Outlook créés vers contacts prestataire\n"
     "2. Docs attendus : bilan PCA N-1 · preuve de réalisation · plan d'action si anomalies\n"
     "3. À réception : saisir Oui/Non col J + date test col K + anomalies col L\n"
     "4. Relances auto via [RELANCES L1-09]"),
    ("ACTIVATION MACROS VBA — une seule fois",
     "1. Ouvrir ce fichier dans Excel\n"
     "2. Alt+F11 → Insertion → Module\n"
     "3. Coller le contenu de MailsLOD1.bas (ou Fichier → Importer)\n"
     "4. Adapter EXPEDITEUR_NOM et EXPEDITEUR_EMAIL en haut du module\n"
     "5. Fermer VBE → Fichier → Enregistrer sous → Classeur Excel avec macros (.xlsm)\n"
     "6. Dans CONTRÔLES LOD1 : Insertion → Formes → affecter les 4 macros aux boutons"),
]:
    ws_me.merge_cells(f"B{row}:C{row}")
    hdr(ws_me[f"B{row}"], titre, bg=MID_BLUE, sz=9)
    ws_me.row_dimensions[row].height = 22
    row += 1
    ws_me.row_dimensions[row].height = max(18 * contenu.count("\n") + 20, 60)
    ws_me.merge_cells(f"B{row}:C{row}")
    c = ws_me[f"B{row}"]
    c.value = contenu
    c.font = Font(name="Arial", size=9)
    c.fill = PatternFill("solid", fgColor=GREY_BG)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    row += 1

brd(ws_me, 1, row, 2, 3)
ws_me.sheet_properties.tabColor = "FF9900"

# ══════════════════════════════════════════════════════════════════════════════
# Ordre des onglets
# ══════════════════════════════════════════════════════════════════════════════
ordre = ["MODE D'EMPLOI", "DATABASE", "REMÉDIATION", "CONTRÔLES LOD1", "SYNTHÈSE"]
for i, name in enumerate(ordre):
    wb.move_sheet(name, offset=wb.index(wb[name]) * -1 + i)

# ══════════════════════════════════════════════════════════════════════════════
# Injection onglet SCRIPTS_VBA caché
# ══════════════════════════════════════════════════════════════════════════════
vba_path = Path("MailsLOD1.bas")
if vba_path.exists():
    VBA_CODE = vba_path.read_text(encoding="utf-8")
    ws_vba = wb.create_sheet("SCRIPTS_VBA")
    ws_vba.sheet_state = "hidden"
    ws_vba.sheet_view.showGridLines = False
    ws_vba.column_dimensions["B"].width = 120
    hdr(ws_vba["B1"],
        "CODE VBA — Alt+F11 → Insertion → Module → Coller ce contenu",
        bg=NAVY, sz=10)
    ws_vba.row_dimensions[1].height = 22
    for i, line in enumerate(VBA_CODE.split("\n"), start=2):
        c = ws_vba.cell(row=i, column=2, value=line)
        c.font = Font(name="Courier New", size=8)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws_vba.row_dimensions[i].height = 13

# ── Sauvegarde ───────────────────────────────────────────────────────────────
OUT = Path("Matrice_PCI_Remediation_v3.xlsx")
wb.save(str(OUT))
print(f"✅ Généré : {OUT.name}  ({N} prestataires)")
print("ℹ  Pour activer les macros : Alt+F11 → Module → coller MailsLOD1.bas → Enregistrer en .xlsm")
