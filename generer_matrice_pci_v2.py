# =============================================================================
#  GÉNÉRATEUR v2 — Matrice de Suivi Remédiation PCI
#  Amundi Immobilier — Opérations & Conformité Externalisations
#  Usage  : python generer_matrice_pci_v2.py
#  Sortie : Matrice_PCI_Remediation_v2.xlsm
#  Prérequis : pip install openpyxl
# =============================================================================

import shutil, datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

# ── Palette ────────────────────────────────────────────────────────────────
NAVY      = "001C4B"; BLUE      = "009EE0"; LIGHT_BLUE= "CCE5F3"
MID_BLUE  = "66B3DB"; WHITE     = "FFFFFF"; GREY_BG   = "F5F7FA"
GREEN_LT  = "D6F0DD"; ORANGE_LT = "FFE5B4"; RED_LT    = "FFD7D7"
YELLOW_LT = "FFFDE7"; GREY_BDR  = "BFCAD4"

def hdr(c, t, bg=NAVY, fg=WHITE, sz=10, bold=True):
    c.value=t; c.font=Font(name="Arial",bold=bold,color=fg,size=sz)
    c.fill=PatternFill("solid",fgColor=bg)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)

def sub(c, t, bg=LIGHT_BLUE, fg=NAVY, sz=9):
    c.value=t; c.font=Font(name="Arial",bold=True,color=fg,size=sz)
    c.fill=PatternFill("solid",fgColor=bg)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)

def std(c, v="", sz=9, bold=False, bg=WHITE, center=False, italic=False):
    c.value=v; c.font=Font(name="Arial",size=sz,bold=bold,italic=italic)
    c.fill=PatternFill("solid",fgColor=bg)
    c.alignment=Alignment(horizontal="center" if center else "left",vertical="center",wrap_text=True)

def brd(ws, r1, r2, c1, c2):
    s=Side(style="thin",color=GREY_BDR); b=Border(left=s,right=s,top=s,bottom=s)
    for row in ws.iter_rows(min_row=r1,max_row=r2,min_col=c1,max_col=c2):
        for c in row: c.border=b

def dv(ws, f, sq):
    v=DataValidation(type="list",formula1=f,allow_blank=True,showErrorMessage=True)
    v.sqref=sq; ws.add_data_validation(v)

def rag(ws, col, r1, r2):
    rng=f"{col}{r1}:{col}{r2}"
    ws.conditional_formatting.add(rng,CellIsRule("equal",['"Vert"'],fill=PatternFill("solid",fgColor=GREEN_LT),font=Font(name="Arial",size=9,color="00703C")))
    ws.conditional_formatting.add(rng,CellIsRule("equal",['"Orange"'],fill=PatternFill("solid",fgColor=ORANGE_LT),font=Font(name="Arial",size=9,color="7D4500")))
    ws.conditional_formatting.add(rng,CellIsRule("equal",['"Rouge"'],fill=PatternFill("solid",fgColor=RED_LT),font=Font(name="Arial",size=9,color="C00000")))

def date_mfc(ws, col, r1, r2):
    rng=f"{col}{r1}:{col}{r2}"
    ws.conditional_formatting.add(rng,FormulaRule(formula=[f'AND(ISNUMBER({col}{r1}),{col}{r1}<TODAY())'],fill=PatternFill("solid",fgColor=RED_LT)))
    ws.conditional_formatting.add(rng,FormulaRule(formula=[f'AND(ISNUMBER({col}{r1}),{col}{r1}>=TODAY(),{col}{r1}<=TODAY()+30)'],fill=PatternFill("solid",fgColor=ORANGE_LT)))

TODAY_STR = datetime.date.today().strftime("%d/%m/%Y")

wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# DATABASE — structure simplifiée
# ══════════════════════════════════════════════════════════════════════════════
ws_db = wb.active; ws_db.title = "DATABASE"
ws_db.sheet_view.showGridLines = False

ws_db.merge_cells("A1:I1")
hdr(ws_db["A1"], "DATABASE — Référentiel Prestataires  |  Tout ajout ici alimente automatiquement les autres onglets", sz=11)
ws_db.row_dimensions[1].height = 28

db_cols = [
    ("A","ID Easy",        16),
    ("B","Prestataire",    28),
    ("C","Criticité\n(PCI / PS / PCMG)", 14),
    ("D","Intragroupe\n(Oui / Non)",      13),
    ("E","Date fin\ncontrat",             14),
    ("F","Contact\nMétier",               22),
    ("G","Mail\nMétier",                  30),
    ("H","Contact\nPrestataire",          22),
    ("I","Mail\nPrestataire",             30),
]
for col, title, width in db_cols:
    sub(ws_db[f"{col}2"], title)
    ws_db.column_dimensions[col].width = width
ws_db.row_dimensions[2].height = 34

# 28 prestataires réels
prestataires = [
    ("CTR126384","Ad Astra",           "PCI", "Non"),
    ("CTR068683","AERIUM",             "PS",  "Non"),
    ("TBD",      "Alter Domus Lux",    "PCI", "Non"),
    ("CTR091882","Amundi Asset Management","PCI","Oui"),
    ("CTR068687","BNPP RE",            "PCMG","Non"),
    ("CTR068688","CA Immo",            "PCMG","Oui"),
    ("CTR148741","CA Immo",            "PCI", "Oui"),
    ("CTR068691","CACEIS",             "PCMG","Oui"),
    ("CTR068694","CDC SF",             "PCI", "Non"),
    ("CTR068696","Constructa",         "PS",  "Non"),
    ("CTR091883","CPR ASSET MANAGEMENT","PCI","Oui"),
    ("CTR116798","Deltager",           "PCI", "Oui"),
    ("CTR068698","Dim4",               "PCI", "Non"),
    ("CTR123300","IC",                 "PCI", "Non"),
    ("CTR068706","IDIA",               "PCI", "Oui"),
    ("CTR068714","Mercialys",          "PS",  "Non"),
    ("CTR068716","Nexity",             "PS",  "Oui"),
    ("CTR129527","PACK SOLUTIONS",     "PCI", "Non"),
    ("CTR068724","SCC",                "PS",  "Non"),
    ("CTR068726","SGSS",               "PS",  "Non"),
    ("CTR077606","Sienna GER",         "PCI", "Non"),
    ("CTR068702","Sienna PB",          "PCI", "Non"),
    ("CTR068729","Sinteo",             "PCI", "Non"),
    ("CTR068732","Sudeco",             "PS",  "Non"),
    ("CTR068720","Telamon",            "PS",  "Non"),
    ("CTR068737","Telmma",             "PCI", "Non"),
    ("CTR068690","UPTEVIA",            "PCI", "Oui"),
    ("CTR136718","Yardi",              "PCI", "Non"),
]

for i, (id_easy, nom, crit, intra) in enumerate(prestataires, start=3):
    r = i
    ws_db.row_dimensions[r].height = 20
    bg = WHITE if r % 2 == 0 else GREY_BG
    for j, val in enumerate([id_easy, nom, crit, intra, "", "", "", "", ""], start=1):
        c = ws_db.cell(row=r, column=j, value=val)
        c.font = Font(name="Arial", size=9)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left" if j in [1,2,6,7,8,9] else "center",
                                  vertical="center")
    ws_db[f"E{r}"].number_format = "DD/MM/YYYY"

dv(ws_db, '"PCI,PS,PCMG"',     "C3:C500")
dv(ws_db, '"Oui,Non"',          "D3:D500")
brd(ws_db, 1, 50, 1, 9)
ws_db.freeze_panes = "A3"
ws_db.sheet_properties.tabColor = NAVY

# ══════════════════════════════════════════════════════════════════════════════
# EASY — tous les champs BCE
# ══════════════════════════════════════════════════════════════════════════════
ws_easy = wb.create_sheet("EASY")
ws_easy.sheet_view.showGridLines = False

ws_easy.merge_cells("A1:H1")
hdr(ws_easy["A1"], "COMPLÉTUDE CHAMPS EASY — Registre BCE  |  Suivi semestriel par prestataire", sz=11)
ws_easy.row_dimensions[1].height = 28

# En-têtes colonnes
easy_hdrs = [
    ("A","Code BCE", 9), ("B","Libellé champ (FR)", 52),
    ("C","Périmètre", 20), ("D","Prestataire", 24),
    ("E","S1 — OK ?\n(Oui/Non/N/A)", 14), ("F","Date vérif. S1", 13), ("G","Statut S1", 10),
    ("H","S2 — OK ?\n(Oui/Non/N/A)", 14), ("I","Date vérif. S2", 13), ("J","Statut S2", 10),
    ("K","Commentaire / Valeur", 34),
]
for col, title, width in easy_hdrs:
    sub(ws_easy[f"{col}2"], title)
    ws_easy.column_dimensions[col].width = width
ws_easy.row_dimensions[2].height = 36

# Champs BCE structurés
# (code_bce, libelle_fr, perimetre)
# Périmètre : "Tous" | "PCI" | "PCI HG" | "DORA"
easy_fields = [
    # ── Bloc : For all outsourcing arrangements ─────────────────────────────
    ("B1","TOUS - Identification",""),
    ("010","Numéro de référence interne","Tous"),
    ("020","Code LEI — Entité signataire du contrat","Tous"),
    ("030","Nom — Entité signataire","Tous"),
    ("040","Pays — Entité signataire","Tous"),
    ("050","Code LEI — Entité supervisée couverte par les services","Tous"),
    ("060","Nom — Entité supervisée","Tous"),
    ("070","Pays — Entité supervisée","Tous"),
    ("B2","TOUS - Détails prestataire",""),
    ("080","Code LEI — Prestataire","Tous"),
    ("090","Numéro d'immatriculation — Prestataire","Tous"),
    ("100","Nom — Prestataire","Tous"),
    ("110","Pays du siège social — Prestataire","Tous"),
    ("120","Adresse enregistrée — Prestataire","Tous"),
    ("130","Externalisation intragroupe (Oui/Non)","Tous"),
    ("140","Nom de l'autorité de supervision du prestataire","Tous"),
    ("145","Code LEI — Société mère du prestataire","Tous"),
    ("146","ID national — Société mère du prestataire","Tous"),
    ("150","Nom — Société mère du prestataire","Tous"),
    ("160","Pays du siège — Société mère","Tous"),
    ("B3","TOUS - Contrat",""),
    ("170","Date de début du contrat","Tous"),
    ("180","Date de prochain renouvellement","Tous"),
    ("190","Date de fin du contrat","Tous"),
    ("200","Préavis de résiliation — Institution","Tous"),
    ("210","Préavis de résiliation — Prestataire","Tous"),
    ("B4","TOUS - Fonction externalisée",""),
    ("220","Catégorie de la fonction externalisée","Tous"),
    ("221","Catégorie niveau 1","Tous"),
    ("222","Catégorie niveau 2","Tous"),
    ("223","Qualification du niveau d'autorisation","Tous"),
    ("230","Description de la fonction externalisée","Tous"),
    ("240","Transfert / traitement de données personnelles","Tous"),
    ("250","Localisation des données","Tous"),
    ("260","Pays de fourniture des services","Tous"),
    # ── Bloc : PCI uniquement ───────────────────────────────────────────────
    ("B5","PCI - Criticité et évaluation",""),
    ("270","Critique ou Important (Oui/Non)","Tous"),
    ("280","Raisons de criticité ou d'importance","PCI"),
    ("290","Date de dernière évaluation de criticité","PCI"),
    ("300","Modèle(s) de service cloud","PCI"),
    ("310","Modèle(s) de déploiement cloud","PCI"),
    ("320","Cloud — nature des données hébergées","PCI"),
    ("330","Conformité aux EBA/GL/2019/02 et lois nationales","PCI"),
    ("340","Coût budgété annuel estimé (€)","PCI"),
    ("350","Date du dernier assessment des risques","PCI"),
    ("360","Synthèse des principaux résultats du risk assessment","PCI"),
    ("370","Organe décisionnel ayant approuvé","PCI"),
    ("380","Droit applicable au contrat d'externalisation","PCI"),
    ("390","Date du dernier audit","PCI"),
    ("400","Date du prochain audit","PCI"),
    ("410","Possibilité de sous-traitance matérielle dans le contrat","PCI"),
    ("415","Code LEI — Sous-traitants","PCI"),
    ("416","ID national — Sous-traitants","PCI"),
    ("417","Sous-traitant appartient au groupe de l'entité signataire","PCI"),
    ("420","Nom(s) — Sous-traitants","PCI"),
    ("430","Pays d'immatriculation — Sous-traitants","PCI"),
    ("440","Pays d'exécution des services — Sous-traitants","PCI"),
    ("450","Localisation des données — Sous-traitants","PCI"),
    ("460","Transfert/traitement données personnelles — Sous-traitants","PCI"),
    ("B6","PCI - Stratégie de sortie",""),
    ("470","Résultat de l'évaluation de substitutabilité du prestataire","PCI"),
    ("480","Possibilité de réintégration de la fonction externalisée","PCI"),
    ("490","Impact de l'interruption de la fonction externalisée","PCI"),
    ("495","Code LEI — Prestataires alternatifs","PCI"),
    ("496","ID national — Prestataires alternatifs","PCI"),
    ("500","Prestataires alternatifs","PCI"),
    ("510","La fonction externalisée soutient des opérations critiques dans le temps","PCI"),
    # ── Bloc : Champs complémentaires Amundi ────────────────────────────────
    ("B7","Champs complémentaires Amundi",""),
    ("SRB","SRB (Oui / Non)","Tous"),
    ("DORA","DORA (Oui / Non)","Tous"),
    ("DO","Dossier d'Opportunité déposé ? (Oui / Non)","Tous"),
    ("QECI","QECI déposé ? (Oui / Non)","Tous"),
    ("SS","Stratégie de sortie déposée ? (Oui / Non)","PCI HG"),
    ("AR","Avis Risque déposé ? (Oui / Non)","PCI"),
    ("IMAS","Déclaration IMAS déposée ? (Oui / Non)","PCI"),
    ("IDC","ID du contrat d'externalisation","Tous"),
    ("PMG","Prestation Majeure de Niveau Groupe","Tous"),
    ("PPCI","PCI de place","PCI"),
    ("SUPID","Référence Easy du fournisseur (SUPID)","Tous"),
    ("FCI","FCI (b_06.01.0050)","PCI"),
]

# Couleurs par bloc
bloc_bg = {
    "B1":"EAF4FB","B2":"FFF3E0","B3":"F0FFF0","B4":"FFFDE7",
    "B5":"F5F0FF","B6":"FFF0F0","B7":"E8F5E9",
}
current_bg = WHITE
row = 3
for item in easy_fields:
    code, libelle, perim = item
    if code.startswith("B"):
        # ligne séparatrice de bloc
        current_bg = bloc_bg.get(code, WHITE)
        ws_easy.row_dimensions[row].height = 18
        ws_easy.merge_cells(f"A{row}:K{row}")
        c = ws_easy[f"A{row}"]
        c.value = f"  {libelle}"
        c.font = Font(name="Arial", size=9, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=MID_BLUE)
        c.alignment = Alignment(horizontal="left", vertical="center")
        row += 1
        continue

    ws_easy.row_dimensions[row].height = 18
    bg = current_bg
    for col, val in [("A",code),("B",libelle),("C",perim),("D","")]:
        c = ws_easy[f"{col}{row}"]
        c.value = val
        c.font = Font(name="Arial", size=9, bold=(col=="A"))
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left" if col in "ABD" else "center",
                                  vertical="center", wrap_text=(col=="B"))
    for col in "EFGHIJK":
        c = ws_easy[f"{col}{row}"]
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws_easy[f"G{row}"] = f'=IF(E{row}="","—",IF(E{row}="Oui","Vert",IF(E{row}="N/A","N/A","Rouge")))'
    ws_easy[f"J{row}"] = f'=IF(H{row}="","—",IF(H{row}="Oui","Vert",IF(H{row}="N/A","N/A","Rouge")))'
    row += 1

dv(ws_easy, '"Oui,Non,N/A"', f"E3:E{row}")
dv(ws_easy, '"Oui,Non,N/A"', f"H3:H{row}")
for col in ["F","I"]:
    for r in range(3, row):
        ws_easy[f"{col}{r}"].number_format = "DD/MM/YYYY"
rag(ws_easy, "G", 3, row); rag(ws_easy, "J", 3, row)
brd(ws_easy, 2, row, 1, 11)
ws_easy.freeze_panes = "D3"
ws_easy.sheet_properties.tabColor = BLUE

# ══════════════════════════════════════════════════════════════════════════════
# DOCUMENTS
# ══════════════════════════════════════════════════════════════════════════════
ws_doc = wb.create_sheet("DOCUMENTS")
ws_doc.sheet_view.showGridLines = False
ws_doc.merge_cells("A1:V1")
hdr(ws_doc["A1"], "COMPLÉTUDE DOCUMENTAIRE — Dossier d'Externalisation", sz=11)
ws_doc.row_dimensions[1].height = 26

for rng, txt, bg in [
    ("A2:C2","IDENTIFICATION",NAVY),
    ("D2:H2","DOSSIER D'OPPORTUNITÉ",MID_BLUE),
    ("I2:M2","QECI",MID_BLUE),
    ("N2:Q2","AVIS RISQUES (PCI)",MID_BLUE),
    ("R2:V2","STRATÉGIE DE SORTIE (PCI HG)",MID_BLUE),
]:
    ws_doc.merge_cells(rng); hdr(ws_doc[rng.split(":")[0]], txt, bg=bg, sz=9)
ws_doc.row_dimensions[2].height = 18

doc_cols = [
    ("A","Prestataire",26),("B","Criticité",10),("C","Intra/HG",10),
    ("D","Requis ?",11),("E","Présent ?",10),("F","Date",12),("G","Lien GED",22),("H","Statut",10),
    ("I","Requis ?",11),("J","Version",12),("K","Date QECI",12),("L","Lien GED",22),("M","Statut",10),
    ("N","Requis ?",11),("O","Rendu ?",10),("P","Date",12),("Q","Statut",10),
    ("R","Requis ?",11),("S","Présente ?",10),("T","Date",12),("U","Lien GED",22),("V","Statut",10),
]
for col, title, width in doc_cols:
    sub(ws_doc[f"{col}3"], title)
    ws_doc.column_dimensions[col].width = width
ws_doc.row_dimensions[3].height = 30

ws_doc.merge_cells("A4:V4")
c = ws_doc["A4"]
c.value = "ℹ  DO = N/A prestataires existants / Obligatoire nouveaux  |  QECI = Tous  |  Avis Risques = PCI  |  Stratégie de Sortie = PCI Hors Groupe"
c.font = Font(name="Arial", size=8, italic=True, color=NAVY)
c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_doc.row_dimensions[4].height = 20

for i in range(len(prestataires)):
    r = i + 5; db_r = i + 3
    bg = WHITE if r % 2 == 0 else GREY_BG
    ws_doc.row_dimensions[r].height = 20
    ws_doc[f"A{r}"] = f'=IFERROR(DATABASE!B{db_r},"")'
    ws_doc[f"B{r}"] = f'=IFERROR(DATABASE!C{db_r},"")'
    ws_doc[f"C{r}"] = f'=IFERROR(DATABASE!D{db_r},"")'
    ws_doc[f"D{r}"] = f'=IF(A{r}="","",IF(B{r}="PCI",IF(ISNUMBER(DATABASE!E{db_r}),"N/A (existant)","Obligatoire"),"Non requis"))'
    ws_doc[f"H{r}"] = f'=IF(A{r}="","",IF(OR(D{r}="N/A (existant)",D{r}="Non requis"),"N/A",IF(AND(E{r}="Oui",F{r}<>""),"Vert",IF(E{r}="Oui","Orange","Rouge"))))'
    ws_doc[f"I{r}"] = f'=IF(A{r}="","","Obligatoire")'
    ws_doc[f"M{r}"] = f'=IF(A{r}="","",IF(AND(J{r}<>"",K{r}<>""),"Vert",IF(J{r}<>"","Orange","Rouge")))'
    ws_doc[f"N{r}"] = f'=IF(A{r}="","",IF(B{r}="PCI","Obligatoire","Non requis"))'
    ws_doc[f"Q{r}"] = f'=IF(A{r}="","",IF(N{r}="Non requis","N/A",IF(AND(O{r}="Oui",P{r}<>""),"Vert",IF(O{r}="Oui","Orange","Rouge"))))'
    ws_doc[f"R{r}"] = f'=IF(A{r}="","",IF(AND(B{r}="PCI",C{r}="Non"),"Obligatoire","Non requis"))'
    ws_doc[f"V{r}"] = f'=IF(A{r}="","",IF(R{r}="Non requis","N/A",IF(AND(S{r}="Oui",T{r}<>""),"Vert",IF(S{r}="Oui","Orange","Rouge"))))'
    for col in "ABCDEFGHIJKLMNOPQRSTUV":
        c = ws_doc[f"{col}{r}"]
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(horizontal="center" if col not in "AGLU" else "left",
                                  vertical="center", wrap_text=True)

for col in ["E","O","S"]: dv(ws_doc, '"Oui,Non,En cours"', f"{col}5:{col}200")
for col in ["H","M","Q","V"]: rag(ws_doc, col, 5, 5+len(prestataires))
brd(ws_doc, 1, 50, 1, 22)
ws_doc.freeze_panes = "A5"
ws_doc.sheet_properties.tabColor = MID_BLUE

# ══════════════════════════════════════════════════════════════════════════════
# CONTRÔLES LOD1
# ══════════════════════════════════════════════════════════════════════════════
ws_ctrl = wb.create_sheet("CONTRÔLES LOD1")
ws_ctrl.sheet_view.showGridLines = False
ws_ctrl.merge_cells("A1:AB1")
hdr(ws_ctrl["A1"], "CONTRÔLES LOD1 — Suivi Périodique par Prestataire PCI", sz=11)
ws_ctrl.row_dimensions[1].height = 26

ws_ctrl.merge_cells("A2:AB2")
c = ws_ctrl["A2"]
c.value = "▶  Macros VBA disponibles :  [📧 LANCER CAMPAGNE L1-08]  [📧 RELANCES L1-08]  [📧 LANCER CAMPAGNE L1-09]  [📧 RELANCES L1-09]  — Créent des brouillons Outlook pour les non-répondants"
c.font = Font(name="Arial", size=9, italic=True, color=NAVY)
c.fill = PatternFill("solid", fgColor=YELLOW_LT)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_ctrl.row_dimensions[2].height = 22

for rng, txt, bg in [
    ("A3:C3","IDENTIFICATION",NAVY),
    ("D3:G3","L1-06 | EASY Entité (Trim.)",MID_BLUE),
    ("H3:K3","L1-07 | EASY Groupe (Remontée)",MID_BLUE),
    ("L3:Q3","L1-08 | Événements Majeurs (Semestriel)",NAVY),
    ("R3:V3","L1-09 | Tests PCA (Annuel T1)",MID_BLUE),
    ("W3:Z3","L1-03/05 | Contrats & Clauses (Semestriel)",MID_BLUE),
    ("AA3:AB3","L1-12 | Audits (Annuel)",MID_BLUE),
]:
    ws_ctrl.merge_cells(rng); hdr(ws_ctrl[rng.split(":")[0]], txt, bg=bg, sz=9)
ws_ctrl.row_dimensions[3].height = 20

ctrl_cols = [
    ("A","Prestataire",26),("B","Criticité",9),("C","Intra/HG",9),
    ("D","Dernière vérif.\nEASY entité",14),("E","Complétude\n(%)",11),("F","Statut L1-06",10),("G","⏱ Prochain",13),
    ("H","Dernière\nremontée",14),("I","Statut L1-07",10),("J","Actions",18),("K","⏱ Prochain",13),
    ("L","S1 — Reporting\nreçu ?",12),("M","Statut S1",9),("N","Commentaire S1",24),
    ("O","S2 — Reporting\nreçu ?",12),("P","Statut S2",9),("Q","Commentaire S2",24),
    ("R","Bilan PCA\nreçu ?",12),("S","Date test",13),("T","Anomalies ?",10),("U","Statut L1-09",10),("V","⏱ Prochain",13),
    ("W","Contrat + clauses\nOK ?",14),("X","Statut",10),("Y","→ DOCS",9),("Z","⏱ Prochain",13),
    ("AA","Date dernier\naudit",14),("AB","⏱ Prochain\naudit",13),
]
for col, title, width in ctrl_cols:
    sub(ws_ctrl[f"{col}4"], title); ws_ctrl.column_dimensions[col].width = width
ws_ctrl.row_dimensions[4].height = 36

n = len(prestataires)
for i in range(n):
    r = i + 5; db_r = i + 3
    bg = WHITE if r % 2 == 0 else GREY_BG
    ws_ctrl.row_dimensions[r].height = 22
    ws_ctrl[f"A{r}"] = f'=IFERROR(DATABASE!B{db_r},"")'
    ws_ctrl[f"B{r}"] = f'=IFERROR(DATABASE!C{db_r},"")'
    ws_ctrl[f"C{r}"] = f'=IFERROR(DATABASE!D{db_r},"")'
    ws_ctrl[f"F{r}"] = f'=IF(A{r}="","",IF(E{r}="","—",IF(E{r}>=95,"Vert",IF(E{r}>=90,"Orange","Rouge"))))'
    ws_ctrl[f"G{r}"] = f'=IF(D{r}="","À planifier",D{r}+92)'
    ws_ctrl[f"I{r}"] = f'=IF(A{r}="","",IF(H{r}="","—",IF(TODAY()-H{r}<=90,"Vert",IF(TODAY()-H{r}<=120,"Orange","Rouge"))))'
    ws_ctrl[f"K{r}"] = f'=IF(H{r}="","À planifier",H{r}+90)'
    ws_ctrl[f"M{r}"] = f'=IF(A{r}="","",IF(L{r}="","—",IF(L{r}="Oui","Vert","Rouge")))'
    ws_ctrl[f"P{r}"] = f'=IF(A{r}="","",IF(O{r}="","—",IF(O{r}="Oui","Vert","Rouge")))'
    ws_ctrl[f"U{r}"] = f'=IF(A{r}="","",IF(R{r}="","—",IF(AND(R{r}="Oui",T{r}="Non"),"Vert",IF(AND(R{r}="Oui",T{r}="Oui"),"Orange","Rouge"))))'
    ws_ctrl[f"V{r}"] = f'=IF(S{r}="","À planifier",DATE(YEAR(S{r})+1,3,31))'
    ws_ctrl[f"X{r}"] = f'=IF(A{r}="","",IF(W{r}="","—",IF(W{r}="Oui","Vert","Rouge")))'
    ws_ctrl[f"Y{r}"] = f'=HYPERLINK("#DOCUMENTS!A1","→ DOCS")'
    ws_ctrl[f"Z{r}"] = f'=IF(W{r}="","À planifier",W{r}+184)'
    ws_ctrl[f"AB{r}"] = f'=IF(AA{r}="","À planifier",DATE(YEAR(AA{r})+1,MONTH(AA{r}),DAY(AA{r})))'
    for col, _, _ in ctrl_cols:
        c = ws_ctrl[f"{col}{r}"]
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.fill = PatternFill("solid", fgColor=bg)

for col in ["F","I","M","P","U","X"]: rag(ws_ctrl, col, 5, 5+n)
for col in ["G","K","V","Z","AB"]:
    date_mfc(ws_ctrl, col, 5, 5+n)
    for r in range(5, 5+n): ws_ctrl[f"{col}{r}"].number_format = "DD/MM/YYYY"
for col in ["D","H","S","AA"]:
    for r in range(5, 5+n): ws_ctrl[f"{col}{r}"].number_format = "DD/MM/YYYY"

dv(ws_ctrl, '"Oui,Non,En attente"', "L5:L500")
dv(ws_ctrl, '"Oui,Non,En attente"', "O5:O500")
dv(ws_ctrl, '"Oui,Non,En cours"',   "R5:R500")
dv(ws_ctrl, '"Oui,Non"',            "T5:T500")
dv(ws_ctrl, '"Oui,Non,En cours"',   "W5:W500")
brd(ws_ctrl, 3, 50, 1, 28)
ws_ctrl.freeze_panes = "A5"
ws_ctrl.sheet_properties.tabColor = NAVY

# ══════════════════════════════════════════════════════════════════════════════
# PS — Prestations Simples
# ══════════════════════════════════════════════════════════════════════════════
ws_ps = wb.create_sheet("PS")
ws_ps.sheet_view.showGridLines = False
ws_ps.merge_cells("A1:L1")
hdr(ws_ps["A1"], "PRESTATIONS SIMPLES (PS) — Registre allégé", sz=11)
ws_ps.row_dimensions[1].height = 26
ps_cols = [
    ("A","ID Easy",14),("B","Prestataire",26),("C","Intragroupe",12),
    ("D","Date fin contrat",14),("E","Contact Métier",22),("F","Mail Métier",28),
    ("G","QECI présent ?",12),("H","DO requis ?",14),("I","DO présent ?",12),
    ("J","Contrat archivé EASY ?",16),("K","Statut global",10),("L","Commentaires",32),
]
for col, title, width in ps_cols:
    sub(ws_ps[f"{col}2"], title); ws_ps.column_dimensions[col].width = width
ws_ps.row_dimensions[2].height = 30

ps_presta = [(id_e, nom, intra) for id_e, nom, crit, intra in prestataires if crit == "PS"]
for i, (id_e, nom, intra) in enumerate(ps_presta, start=3):
    r = i; bg = WHITE if r % 2 == 0 else GREY_BG
    ws_ps.row_dimensions[r].height = 20
    for j, val in enumerate([id_e, nom, intra, "", "", "", "", "", "", "", "", ""], start=1):
        c = ws_ps.cell(row=r, column=j, value=val)
        c.font = Font(name="Arial", size=9)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left" if j in [1,2,5,6,12] else "center", vertical="center")
    ws_ps[f"K{r}"] = f'=IF(B{r}="","",IF(AND(G{r}="Oui",J{r}="Oui"),"Vert",IF(OR(G{r}="Non",J{r}="Non"),"Rouge","Orange")))'
    ws_ps[f"D{r}"].number_format = "DD/MM/YYYY"

dv(ws_ps, '"Oui,Non,En cours"', "G3:G100")
dv(ws_ps, '"Obligatoire (nouveau),N/A (existant)"', "H3:H100")
dv(ws_ps, '"Oui,Non,N/A"', "I3:I100")
dv(ws_ps, '"Oui,Non,En cours"', "J3:J100")
rag(ws_ps, "K", 3, 3+len(ps_presta))
brd(ws_ps, 1, 50, 1, 12)
ws_ps.freeze_panes = "A3"
ws_ps.sheet_properties.tabColor = MID_BLUE

# ══════════════════════════════════════════════════════════════════════════════
# SYNTHÈSE
# ══════════════════════════════════════════════════════════════════════════════
ws_syn = wb.create_sheet("SYNTHÈSE")
ws_syn.sheet_view.showGridLines = False
ws_syn.merge_cells("A1:I1")
hdr(ws_syn["A1"], f"SYNTHÈSE REMÉDIATION PCI — Tableau de bord  |  MàJ : {TODAY_STR}", sz=12)
ws_syn.row_dimensions[1].height = 28
syn_cols = [
    ("A","Prestataire",28),("B","ID Easy",14),("C","Criticité",10),
    ("D","Docs",10),("E","EASY",10),("F","L1-08 S1",10),("G","L1-09 PCA",10),
    ("H","⚠ Risque\nglobal",11),("I","Commentaire / Action",36),
]
for col, title, width in syn_cols:
    sub(ws_syn[f"{col}2"], title); ws_syn.column_dimensions[col].width = width
ws_syn.row_dimensions[2].height = 30

for i in range(n):
    r = i + 3; db_r = i + 3; ctrl_r = i + 5
    bg = WHITE if r % 2 == 0 else GREY_BG
    ws_syn.row_dimensions[r].height = 20
    ws_syn[f"A{r}"] = f'=IFERROR(DATABASE!B{db_r},"")'
    ws_syn[f"B{r}"] = f'=IFERROR(DATABASE!A{db_r},"")'
    ws_syn[f"C{r}"] = f'=IFERROR(DATABASE!C{db_r},"")'
    ws_syn[f"D{r}"] = f'=IFERROR(DOCUMENTS!H{ctrl_r},"—")'
    ws_syn[f"E{r}"] = f'=IFERROR(\'CONTRÔLES LOD1\'!F{ctrl_r},"—")'
    ws_syn[f"F{r}"] = f'=IFERROR(\'CONTRÔLES LOD1\'!M{ctrl_r},"—")'
    ws_syn[f"G{r}"] = f'=IFERROR(\'CONTRÔLES LOD1\'!U{ctrl_r},"—")'
    ws_syn[f"H{r}"] = f'=IF(A{r}="","",IF(OR(D{r}="Rouge",E{r}="Rouge",F{r}="Rouge",G{r}="Rouge"),"Rouge",IF(OR(D{r}="Orange",E{r}="Orange",F{r}="Orange",G{r}="Orange"),"Orange","Vert")))'
    for col, _, _ in syn_cols:
        c = ws_syn[f"{col}{r}"]
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(horizontal="left" if col == "I" else "center", vertical="center")
        c.fill = PatternFill("solid", fgColor=bg)

for col in ["D","E","F","G","H"]: rag(ws_syn, col, 3, 3+n)
brd(ws_syn, 2, 50, 1, 9)
ws_syn.freeze_panes = "A3"
ws_syn.sheet_properties.tabColor = "FF0000"

# ══════════════════════════════════════════════════════════════════════════════
# MODE D'EMPLOI
# ══════════════════════════════════════════════════════════════════════════════
ws_me = wb.create_sheet("MODE D'EMPLOI")
ws_me.sheet_view.showGridLines = False
ws_me.column_dimensions["A"].width = 3
ws_me.column_dimensions["B"].width = 26
ws_me.column_dimensions["C"].width = 62
ws_me.merge_cells("B1:C1")
hdr(ws_me["B1"], f"MODE D'EMPLOI — Matrice Remédiation PCI  |  Amundi Immobilier  |  {TODAY_STR}", sz=11)
ws_me.row_dimensions[1].height = 28

# Onglets
row = 3
ws_me.merge_cells(f"B{row}:C{row}"); hdr(ws_me[f"B{row}"], "STRUCTURE DU FICHIER", bg=NAVY, sz=10)
ws_me.row_dimensions[row].height = 22; row += 1
onglets_desc = [
    ("DATABASE",       "Référentiel simplifié : 9 colonnes (ID Easy, Prestataire, Criticité, Intragroupe, Date fin contrat, Contact Métier, Mail Métier, Contact Prestataire, Mail Prestataire). Saisie unique — alimente tous les autres onglets."),
    ("EASY",           "Tous les champs du registre BCE (codes 010 à FCI) + champs complémentaires Amundi. Suivi semestriel S1/S2 avec statut RAG calculé automatiquement."),
    ("DOCUMENTS",      "DO / QECI / Avis Risques / Stratégie de Sortie. Colonnes Requis calculées automatiquement selon criticité et intra/hors groupe."),
    ("CONTRÔLES LOD1", "Tableau central : L1-06, L1-07, L1-08 (S1+S2), L1-09, L1-03/05, L1-12. Dates prochain contrôle calculées + MFC rouge/orange si dépassées. Boutons macros VBA."),
    ("PS",             "Suivi allégé Prestations Simples (QECI + archivage EASY). Pré-rempli avec les PS de la DATABASE."),
    ("SYNTHÈSE",       "Tableau de bord consolidé RAG. Lecture seule — mis à jour automatiquement."),
]
for nom, desc in onglets_desc:
    ws_me.row_dimensions[row].height = 40
    bg = WHITE if row % 2 == 0 else GREY_BG
    c = ws_me[f"B{row}"]
    c.value = nom; c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    std(ws_me[f"C{row}"], desc, bg=bg, sz=9)
    row += 1

# Procédures
row += 1
for titre, contenu in [
    ("PROCÉDURE L1-08 (Semestriel)",
     "Périmètre : PCI Hors Groupe  |  Destinataires : Mail Métier (col G de DATABASE)\n"
     "S1 : Lancement 1er juin → Relance J-7 (23/06) → J-3 (27/06) → J-1 (29/06) → Deadline 30 juin\n"
     "S2 : Lancement 1er déc  → Relance J-7 (23/12) → J-3 (27/12) → J-1 (29/12) → Deadline 30 déc\n"
     "Bouton Excel : [📧 LANCER CAMPAGNE L1-08] puis [📧 RELANCES L1-08] pour non-répondants"),
    ("PROCÉDURE L1-09 (Annuel T1)",
     "Périmètre : Toutes PCI  |  Destinataires : Mail Prestataire (col I de DATABASE)\n"
     "Lancement 15 janvier → Relance J-7 (24/03) → J-3 (28/03) → J-1 (30/03) → Deadline 31 mars\n"
     "Bouton Excel : [📧 LANCER CAMPAGNE L1-09] puis [📧 RELANCES L1-09] pour non-répondants\n"
     "Documents attendus : bilan tests PCA N-1, preuve de réalisation, plan d'action si anomalies"),
    ("ACTIVATION MACROS VBA (une seule fois)",
     "1. Ouvrir le .xlsm dans Excel\n"
     "2. Alt+F11 → Insertion → Module\n"
     "3. Coller le contenu du fichier MailsLOD1.bas (ou Fichier → Importer)\n"
     "4. Adapter EXPEDITEUR_NOM et EXPEDITEUR_EMAIL en haut du module\n"
     "5. Fermer VBE → Sauvegarder en .xlsm\n"
     "6. Dans CONTRÔLES LOD1 : Insertion → Formes → affecter les 4 macros aux boutons"),
]:
    ws_me.merge_cells(f"B{row}:C{row}")
    hdr(ws_me[f"B{row}"], titre, bg=MID_BLUE, sz=9); ws_me.row_dimensions[row].height = 22; row += 1
    ws_me.row_dimensions[row].height = 80
    ws_me.merge_cells(f"B{row}:C{row}")
    c = ws_me[f"B{row}"]
    c.value = contenu; c.font = Font(name="Arial", size=9)
    c.fill = PatternFill("solid", fgColor=GREY_BG)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    row += 1

brd(ws_me, 1, row, 2, 3)
ws_me.sheet_properties.tabColor = "FF9900"

# ══════════════════════════════════════════════════════════════════════════════
# Réordonnancement onglets
# ══════════════════════════════════════════════════════════════════════════════
ordre = ["MODE D'EMPLOI","DATABASE","DOCUMENTS","EASY","CONTRÔLES LOD1","PS","SYNTHÈSE"]
for i, name in enumerate(ordre):
    wb.move_sheet(name, offset=wb.index(wb[name]) * -1 + i)

# ══════════════════════════════════════════════════════════════════════════════
# Sauvegarde + injection onglet SCRIPTS_VBA caché
# ══════════════════════════════════════════════════════════════════════════════
TEMP = Path("/home/claude/matrice_v2_temp.xlsx")
OUT  = Path("/home/claude/Matrice_PCI_Remediation_v2.xlsm")
wb.save(str(TEMP))

VBA_CODE = open("/home/claude/MailsLOD1.bas", encoding="utf-8").read()
wb2 = load_workbook(str(TEMP))
ws_vba = wb2.create_sheet("SCRIPTS_VBA")
ws_vba.sheet_state = "hidden"
ws_vba.sheet_view.showGridLines = False
ws_vba.column_dimensions["B"].width = 120
hdr(ws_vba["B1"], "CODE VBA — Alt+F11 → Insertion → Module → Coller ce contenu", bg=NAVY, sz=10)
ws_vba.row_dimensions[1].height = 22
for i, line in enumerate(VBA_CODE.split("\n"), start=2):
    c = ws_vba.cell(row=i, column=2, value=line)
    c.font = Font(name="Courier New", size=8)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws_vba.row_dimensions[i].height = 13

wb2.save(str(OUT))
TEMP.unlink(missing_ok=True)
print(f"✅ Généré : {OUT.name}  ({len(prestataires)} prestataires, champs BCE complets)")
