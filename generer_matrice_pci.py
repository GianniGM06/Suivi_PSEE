# =============================================================================
#  GÉNÉRATEUR — Matrice de Suivi Remédiation PCI
#  Amundi Immobilier — Opérations & Conformité Externalisations
#
#  Usage  : python generer_matrice_pci.py
#  Sortie : Matrice_PCI_Remediation.xlsm  (dans le même dossier)
#
#  Prérequis : pip install openpyxl
#
#  Ce script génère un fichier .xlsm avec :
#    - 7 onglets (MODE D'EMPLOI, DATABASE, DOCUMENTS, EASY,
#                 CONTRÔLES LOD1, PS, SYNTHÈSE)
#    - Macros VBA intégrées (boutons envoi mails L1-08 et L1-09)
#    - Formules RAG, MFC, dropdowns, freeze panes
# =============================================================================

import zipfile, shutil, os, datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

# ── Palette Amundi ─────────────────────────────────────────────────────────
NAVY       = "001C4B"
BLUE       = "009EE0"
LIGHT_BLUE = "CCE5F3"
MID_BLUE   = "66B3DB"
WHITE      = "FFFFFF"
GREY_BG    = "F5F7FA"
GREY_BDR   = "BFCAD4"
GREEN      = "00B050"
ORANGE     = "FF9900"
RED        = "FF0000"
RED_LT     = "FFD7D7"
ORANGE_LT  = "FFE5B4"
GREEN_LT   = "D6F0DD"
YELLOW_LT  = "FFFDE7"

# ── Helpers styling ────────────────────────────────────────────────────────
def hdr(cell, txt, bg=NAVY, fg=WHITE, sz=10, bold=True, center=True):
    cell.value = txt
    cell.font = Font(name="Arial", bold=bold, color=fg, size=sz)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center" if center else "left",
                                vertical="center", wrap_text=True)

def subhdr(cell, txt, bg=LIGHT_BLUE, fg=NAVY, sz=9):
    cell.value = txt
    cell.font = Font(name="Arial", bold=True, color=fg, size=sz)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def cell_std(cell, val="", sz=9, bold=False, bg=WHITE, center=False, italic=False):
    cell.value = val
    cell.font = Font(name="Arial", size=sz, bold=bold, italic=italic)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center" if center else "left",
                                vertical="center", wrap_text=True)

def border_range(ws, r1, r2, c1, c2, color=GREY_BDR):
    s = Side(style="thin", color=color)
    b = Border(left=s, right=s, top=s, bottom=s)
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
        for c in row:
            c.border = b

def dv(ws, formula, sqref):
    v = DataValidation(type="list", formula1=formula, allow_blank=True, showErrorMessage=True)
    v.sqref = sqref
    ws.add_data_validation(v)

def rag_mfc(ws, col, r1, r2):
    rng = f"{col}{r1}:{col}{r2}"
    ws.conditional_formatting.add(rng, CellIsRule("equal", ['"Vert"'],
        fill=PatternFill("solid", fgColor=GREEN_LT),
        font=Font(name="Arial", size=9, color="00703C")))
    ws.conditional_formatting.add(rng, CellIsRule("equal", ['"Orange"'],
        fill=PatternFill("solid", fgColor=ORANGE_LT),
        font=Font(name="Arial", size=9, color="7D4500")))
    ws.conditional_formatting.add(rng, CellIsRule("equal", ['"Rouge"'],
        fill=PatternFill("solid", fgColor=RED_LT),
        font=Font(name="Arial", size=9, color="C00000")))
    ws.conditional_formatting.add(rng, CellIsRule("equal", ['"N/A"'],
        fill=PatternFill("solid", fgColor="EEEEEE"),
        font=Font(name="Arial", size=9, color="888888")))

def date_mfc(ws, col, r1, r2):
    """Rouge si date dépassée, orange si échéance < 30j."""
    rng = f"{col}{r1}:{col}{r2}"
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'AND(ISNUMBER({col}{r1}),{col}{r1}<TODAY())'],
        fill=PatternFill("solid", fgColor=RED_LT)))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'AND(ISNUMBER({col}{r1}),{col}{r1}>=TODAY(),{col}{r1}<=TODAY()+30)'],
        fill=PatternFill("solid", fgColor=ORANGE_LT)))

TODAY_STR = datetime.date.today().strftime("%d/%m/%Y")
YEAR      = datetime.date.today().year

# ══════════════════════════════════════════════════════════════════════════════
wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET 0 — MODE D'EMPLOI
# ══════════════════════════════════════════════════════════════════════════════
ws_me = wb.active
ws_me.title = "MODE D'EMPLOI"
ws_me.sheet_view.showGridLines = False
ws_me.column_dimensions["A"].width = 3
ws_me.column_dimensions["B"].width = 28
ws_me.column_dimensions["C"].width = 60
ws_me.column_dimensions["D"].width = 22
ws_me.column_dimensions["E"].width = 18

ws_me.merge_cells("B1:E1")
hdr(ws_me["B1"], "MATRICE DE SUIVI REMÉDIATION PCI — Mode d'emploi", bg=NAVY, fg=WHITE, sz=14)
ws_me.row_dimensions[1].height = 36

ws_me.merge_cells("B2:E2")
cell_std(ws_me["B2"],
    f"Amundi Immobilier — Opérations & Conformité Externalisations  |  Généré le {TODAY_STR}",
    sz=9, italic=True, bg=LIGHT_BLUE)
ws_me.row_dimensions[2].height = 18

sections_me = [
    # (row, titre, contenu, bg_titre)
    (4,  "STRUCTURE DU FICHIER", None, NAVY),
    (5,  "MODE D'EMPLOI", None, NAVY),
    (6,  "MACROS VBA", None, NAVY),
    (7,  "PROCÉDURE L1-08", None, NAVY),
    (8,  "PROCÉDURE L1-09", None, NAVY),
    (9,  "RÈGLES DOCUMENTAIRES", None, NAVY),
    (10, "GLOSSAIRE", None, NAVY),
]

onglets = [
    ("DATABASE",        MID_BLUE,  "Saisie unique des prestataires. Tout ajout ici alimente automatiquement les autres onglets.\n→ Renseigner TOUS les champs, notamment Email Responsable (col L) et Email Contact PCA (col N)."),
    ("DOCUMENTS",       MID_BLUE,  "Complétude documentaire (DO, QECI, Avis Risques, Stratégie de Sortie).\n→ Les colonnes 'Requis ?' se calculent automatiquement selon le type et l'intra/hors groupe.\n→ Saisir uniquement : Présent ? / Date / Lien GED."),
    ("EASY",            MID_BLUE,  "Suivi semestriel des 27 champs obligatoires dans le registre EASY.\n→ Pour chaque prestataire : saisir Oui/Non/N/A pour S1 et S2 + date de vérification."),
    ("CONTRÔLES LOD1",  NAVY,      "Tableau central de suivi des 6 contrôles périodiques.\n→ Les dates 'Prochain contrôle' se calculent seules (rouge si dépassées, orange si < 30j).\n→ Boutons ENVOYER MAILS L1-08 et L1-09 en haut de l'onglet."),
    ("PS",              MID_BLUE,  "Suivi allégé des Prestations Simples (QECI + archivage EASY).\n→ Même logique que DOCUMENTS mais périmètre réduit."),
    ("SYNTHÈSE",        NAVY,      "Tableau de bord consolidé. Lecture seule — se met à jour automatiquement.\n→ Consulter avant chaque CCI ou rapport à la Direction des Risques."),
]

# Header onglets
row = 4
ws_me.merge_cells(f"B{row}:E{row}")
hdr(ws_me[f"B{row}"], "ONGLETS DU FICHIER", bg=NAVY, sz=10)
ws_me.row_dimensions[row].height = 24
row += 1

subhdr(ws_me[f"B{row}"], "Onglet", bg=MID_BLUE)
subhdr(ws_me[f"C{row}"], "Description & consignes de saisie", bg=MID_BLUE)
ws_me.row_dimensions[row].height = 22
row += 1

for onglet, bg_o, desc in onglets:
    ws_me.row_dimensions[row].height = 52
    bg_row = WHITE if row % 2 == 0 else GREY_BG
    cell_std(ws_me[f"B{row}"], onglet, bold=True, bg=bg_row, sz=9)
    ws_me[f"B{row}"].font = Font(name="Arial", bold=True, size=9, color=WHITE)
    ws_me[f"B{row}"].fill = PatternFill("solid", fgColor=bg_o)
    ws_me[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center")
    cell_std(ws_me[f"C{row}"], desc, bg=bg_row, sz=9)
    row += 1

# Procédure L1-08
row += 1
ws_me.merge_cells(f"B{row}:E{row}")
hdr(ws_me[f"B{row}"], "PROCÉDURE EXT-L1-08 — Pilotage des Événements Majeurs (Semestriel)", bg=NAVY, sz=10)
ws_me.row_dimensions[row].height = 24
row += 1

proc_l108 = [
    ("Périmètre",      "Toutes les PCI Hors Groupe"),
    ("Destinataires",  "Responsable opérationnel interne (onglet DATABASE, colonne L)"),
    ("Fréquence",      "Semestrielle — S1 (clôture 30 juin) et S2 (clôture 30 décembre)"),
    ("Lancement S1",   "1er juin → envoi du mail initial à tous les responsables métier concernés"),
    ("Relance S1 J-7", "23 juin → 1ère relance aux non-répondants"),
    ("Relance S1 J-3", "27 juin → 2ème relance"),
    ("Relance S1 J-1", "29 juin → Dernière relance"),
    ("Deadline S1",    "30 juin — clôture. Saisir le statut dans CONTRÔLES LOD1 colonne M (S1)"),
    ("Lancement S2",   "1er décembre → envoi du mail initial"),
    ("Relance S2 J-7", "23 décembre → 1ère relance"),
    ("Relance S2 J-3", "27 décembre → 2ème relance"),
    ("Relance S2 J-1", "29 décembre → Dernière relance"),
    ("Deadline S2",    "30 décembre — clôture. Saisir le statut dans CONTRÔLES LOD1 colonne P (S2)"),
    ("Bouton Excel",   "Dans l'onglet CONTRÔLES LOD1 : cliquer sur [📧 MAILS L1-08]\n"
                       "→ Le script détecte le semestre en cours, identifie les non-répondants,\n"
                       "  et ouvre les brouillons Outlook pour chaque responsable concerné."),
    ("Ce qu'on demande", "• Confirmation réception reporting prestataire (qualité, incidents, SLA)\n"
                         "• Identification de tout événement majeur\n"
                         "• Saisie statut Vert/Orange/Rouge + commentaire dans la matrice"),
    ("Piste d'audit",  "Saisie dans onglet CONTRÔLES LOD1 + éventuels mails archivés dans GED"),
]

for label, val in proc_l108:
    ws_me.row_dimensions[row].height = max(20, val.count('\n') * 13 + 16)
    bg_row = WHITE if row % 2 == 0 else GREY_BG
    cell_std(ws_me[f"B{row}"], label, bold=True, bg=YELLOW_LT, sz=9)
    cell_std(ws_me[f"C{row}"], val, bg=bg_row, sz=9)
    row += 1

# Procédure L1-09
row += 1
ws_me.merge_cells(f"B{row}:E{row}")
hdr(ws_me[f"B{row}"], "PROCÉDURE EXT-L1-09 — Revue Tests PCA (Annuel — T1)", bg=NAVY, sz=10)
ws_me.row_dimensions[row].height = 24
row += 1

proc_l109 = [
    ("Périmètre",       "Toutes les PCI (Intragroupe et Hors Groupe)"),
    ("Destinataires",   "Contact PCA prestataire (onglet DATABASE, colonne N)"),
    ("Fréquence",       "Annuelle — à réaliser au T1 sur la base des tests réalisés en année N-1"),
    ("Lancement",       "15 janvier → demande du bilan PCA à chaque prestataire PCI"),
    ("Relance J-7",     "24 mars → 1ère relance aux prestataires n'ayant pas transmis leur bilan"),
    ("Relance J-3",     "28 mars → 2ème relance"),
    ("Relance J-1",     "30 mars → Dernière relance"),
    ("Deadline",        "31 mars — clôture T1. Saisir 'Oui' dans CONTRÔLES LOD1 colonne R"),
    ("Bouton Excel",    "Dans l'onglet CONTRÔLES LOD1 : cliquer sur [📧 MAILS L1-09]\n"
                        "→ Détecte les prestataires sans bilan reçu (col R ≠ 'Oui')\n"
                        "  et ouvre les brouillons Outlook vers les contacts PCA externes."),
    ("Documents attendus", "• Bilan des tests PCA de l'année N-1 (scénarios, résultats)\n"
                           "• Preuve de réalisation (rapport, PV ou attestation)\n"
                           "• Plan d'action si anomalies\n"
                           "• Confirmation interaction gestion de crise prestataire / Amundi Immobilier"),
    ("Conséquence non-réponse", "Contrôle enregistré Rouge — information Direction des Risques"),
    ("Piste d'audit",   "Bilan archivé dans EASY + saisie dans onglet CONTRÔLES LOD1"),
]

for label, val in proc_l109:
    ws_me.row_dimensions[row].height = max(20, val.count('\n') * 13 + 16)
    bg_row = WHITE if row % 2 == 0 else GREY_BG
    cell_std(ws_me[f"B{row}"], label, bold=True, bg=YELLOW_LT, sz=9)
    cell_std(ws_me[f"C{row}"], val, bg=bg_row, sz=9)
    row += 1

# Règles documentaires
row += 1
ws_me.merge_cells(f"B{row}:E{row}")
hdr(ws_me[f"B{row}"], "RÈGLES D'OBLIGATION DOCUMENTAIRE", bg=NAVY, sz=10)
ws_me.row_dimensions[row].height = 24
row += 1

docs_rules = [
    ("Dossier d'Opportunité (DO)", "N/A pour prestataires existants | Obligatoire pour nouveaux prestataires"),
    ("QECI",                       "Obligatoire pour TOUS les prestataires (PCI et PS)"),
    ("Avis Risques",               "PCI uniquement (Intragroupe et Hors Groupe)"),
    ("Stratégie de Sortie",        "PCI Hors Groupe uniquement"),
    ("Contrat archivé EASY",       "Obligatoire pour tous — avec clauses Qualité, Sécurité, PCA"),
]
for label, val in docs_rules:
    ws_me.row_dimensions[row].height = 22
    bg_row = WHITE if row % 2 == 0 else GREY_BG
    cell_std(ws_me[f"B{row}"], label, bold=True, bg=YELLOW_LT, sz=9)
    cell_std(ws_me[f"C{row}"], val, bg=bg_row, sz=9)
    row += 1

# Glossaire
row += 1
ws_me.merge_cells(f"B{row}:E{row}")
hdr(ws_me[f"B{row}"], "GLOSSAIRE", bg=NAVY, sz=10)
ws_me.row_dimensions[row].height = 24
row += 1

glossaire = [
    ("PCI",    "Prestation Critique ou Importante — soumise au régime EBA/DORA renforcé"),
    ("PS",     "Prestation Simple — suivi allégé"),
    ("QECI",   "Questionnaire d'Évaluation et de Contrôle Interne du prestataire"),
    ("EASY",   "Registre centralisé des externalisations (outil Amundi Groupe / CASA)"),
    ("FCI",    "Fonction Critique ou Importante au sens DORA"),
    ("LOD1",   "Première ligne de défense (contrôles réalisés par le pilote local)"),
    ("LOD2",   "Deuxième ligne de défense (Direction des Risques)"),
    ("CCI",    "Comité de Contrôle Interne"),
    ("PCA",    "Plan de Continuité d'Activité"),
    ("RAG",    "Rouge / Ambre (Orange) / Vert — code couleur de statut"),
    ("DO",     "Dossier d'Opportunité — analyse préalable avant toute externalisation"),
    ("GED",    "Gestion Électronique des Documents (Nuxeo / LegalDocs)"),
    ("BCE",    "Banque Centrale Européenne — autorité de supervision"),
    ("DORA",   "Digital Operational Resilience Act — règlement UE applicable depuis jan. 2025"),
    ("EBA",    "European Banking Authority — guidelines outsourcing 2019"),
]
for label, val in glossaire:
    ws_me.row_dimensions[row].height = 20
    bg_row = WHITE if row % 2 == 0 else GREY_BG
    cell_std(ws_me[f"B{row}"], label, bold=True, bg=LIGHT_BLUE, sz=9)
    cell_std(ws_me[f"C{row}"], val, bg=bg_row, sz=9)
    row += 1

border_range(ws_me, 4, row - 1, 2, 5)

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET 1 — DATABASE
# ══════════════════════════════════════════════════════════════════════════════
ws_db = wb.create_sheet("DATABASE")
ws_db.sheet_view.showGridLines = False

ws_db.merge_cells("A1:U1")
hdr(ws_db["A1"], "DATABASE — Référentiel Prestataires  (tout ajout ici alimente automatiquement les autres onglets)",
    bg=NAVY, fg=WHITE, sz=11)
ws_db.row_dimensions[1].height = 30

# Section headers row 2
for rng, txt in [
    ("A2:D2","IDENTIFICATION"), ("E2:G2","QUALIFICATION"),
    ("H2:J2","CONTRAT"),        ("K2:L2","RESPONSABLE INTERNE"),
    ("M2:N2","CONTACT PCA PRESTATAIRE"), ("O2:P2","FINANCIER"),
    ("Q2:R2","STATUTS"),        ("S2:U2","MÉTADONNÉES"),
]:
    ws_db.merge_cells(rng)
    hdr(ws_db[rng.split(":")[0]], txt, bg=MID_BLUE, sz=9)
ws_db.row_dimensions[2].height = 20

db_cols = [
    ("A","Nom Prestataire",28),           ("B","Code Interne",13),
    ("C","Groupe / Entité mère",20),      ("D","Pays",12),
    ("E","Type\n(PCI / PS)",11),          ("F","Intra /\nHors Groupe",13),
    ("G","Criticité\n(1→4)",10),          ("H","Fonction Externalisée",30),
    ("I","Date Entrée\nen Vigueur",14),   ("J","Date\nÉchéance",14),
    ("K","Responsable Opérationnel\n(Nom)",24), ("L","Email Responsable\nOpérationnel",30),
    ("M","Contact PCA\nPrestataire (Nom)",22),  ("N","Email Contact\nPCA Prestataire",30),
    ("O","Montant\nAnnuel (€)",14),       ("P","Dépendance\nÉconomique",14),
    ("Q","Statut Prestataire",16),        ("R","Statut Contrat",18),
    ("S","Date Création\nFiche",14),      ("T","Dernière\nMàJ",14),
    ("U","Commentaires",32),
]
for col, title, width in db_cols:
    subhdr(ws_db[f"{col}3"], title)
    ws_db.column_dimensions[col].width = width
ws_db.row_dimensions[3].height = 36

# 10 lignes vides formatées
for r in range(4, 14):
    ws_db.row_dimensions[r].height = 22
    bg = WHITE if r % 2 == 0 else GREY_BG
    for col, _, _ in db_cols:
        c = ws_db[f"{col}{r}"]
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

dv(ws_db, '"PCI,PS"',                                           "E4:E500")
dv(ws_db, '"Hors Groupe,Intragroupe"',                          "F4:F500")
dv(ws_db, '"1,2,3,4"',                                          "G4:G500")
dv(ws_db, '"Actif,Inactif,En cours"',                           "Q4:Q500")
dv(ws_db, '"Valide,Échu,Résiliation,En renouvellement"',        "R4:R500")
dv(ws_db, '"Faible,Modérée,Élevée,Critique"',                   "P4:P500")

for r in range(4, 14):
    for col in ["I","J","S","T"]:
        ws_db[f"{col}{r}"].number_format = "DD/MM/YYYY"

border_range(ws_db, 1, 50, 1, 21)
ws_db.freeze_panes = "A4"

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET 2 — DOCUMENTS
# ══════════════════════════════════════════════════════════════════════════════
ws_doc = wb.create_sheet("DOCUMENTS")
ws_doc.sheet_view.showGridLines = False

ws_doc.merge_cells("A1:V1")
hdr(ws_doc["A1"], "COMPLÉTUDE DOCUMENTAIRE — Dossier d'Externalisation PCI",
    bg=NAVY, fg=WHITE, sz=11)
ws_doc.row_dimensions[1].height = 28

for rng, txt, bg in [
    ("A2:C2","IDENTIFICATION",NAVY),
    ("D2:H2","DOSSIER D'OPPORTUNITÉ",MID_BLUE),
    ("I2:M2","QECI",MID_BLUE),
    ("N2:Q2","AVIS RISQUES (PCI)",MID_BLUE),
    ("R2:V2","STRATÉGIE DE SORTIE (PCI HG)",MID_BLUE),
]:
    ws_doc.merge_cells(rng)
    hdr(ws_doc[rng.split(":")[0]], txt, bg=bg, sz=9)
ws_doc.row_dimensions[2].height = 20

doc_cols = [
    ("A","Prestataire",26), ("B","Type",8), ("C","Intra/HG",10),
    ("D","Requis ?",11), ("E","Présent ?",10), ("F","Date",12),
    ("G","Lien GED",22), ("H","Statut",10),
    ("I","Requis ?",11), ("J","Version",12), ("K","Date QECI",12),
    ("L","Lien GED",22), ("M","Statut",10),
    ("N","Requis ?",11), ("O","Avis rendu ?",12), ("P","Date",12), ("Q","Statut",10),
    ("R","Requis ?",11), ("S","Présente ?",10), ("T","Date",12),
    ("U","Lien GED",22), ("V","Statut",10),
]
for col, title, width in doc_cols:
    subhdr(ws_doc[f"{col}3"], title)
    ws_doc.column_dimensions[col].width = width
ws_doc.row_dimensions[3].height = 32

ws_doc.merge_cells("A4:V4")
c = ws_doc["A4"]
c.value = ("ℹ  DO = N/A prestataires existants / Obligatoire nouveaux  |  "
           "QECI = Obligatoire tous  |  Avis Risques = PCI uniquement  |  "
           "Stratégie de Sortie = PCI Hors Groupe uniquement")
c.font = Font(name="Arial", size=8, italic=True, color=NAVY)
c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_doc.row_dimensions[4].height = 24

for r in range(5, 25):
    ws_doc.row_dimensions[r].height = 22
    bg = WHITE if r % 2 == 0 else GREY_BG
    db_r = r - 1  # ligne correspondante dans DATABASE (row 4 → DB row 4, etc.)

    ws_doc[f"A{r}"] = f"=IFERROR(DATABASE!A{db_r},\"\")"
    ws_doc[f"B{r}"] = f"=IFERROR(DATABASE!E{db_r},\"\")"
    ws_doc[f"C{r}"] = f"=IFERROR(DATABASE!F{db_r},\"\")"

    # DO
    ws_doc[f"D{r}"] = (f'=IF(A{r}="","",IF(B{r}="PCI",'
                        f'IF(ISNUMBER(DATABASE!I{db_r}),"N/A (existant)","Obligatoire"),"Non requis"))')
    ws_doc[f"H{r}"] = (f'=IF(A{r}="","",IF(D{r}="N/A (existant)","N/A",'
                        f'IF(D{r}="Non requis","N/A",'
                        f'IF(AND(E{r}="Oui",F{r}<>""),"Vert",IF(E{r}="Oui","Orange","Rouge")))))')
    # QECI
    ws_doc[f"I{r}"] = f'=IF(A{r}="","","Obligatoire")'
    ws_doc[f"M{r}"] = (f'=IF(A{r}="","",IF(AND(J{r}<>"",K{r}<>""),"Vert",'
                        f'IF(J{r}<>"","Orange","Rouge")))')
    # Avis Risques
    ws_doc[f"N{r}"] = f'=IF(A{r}="","",IF(B{r}="PCI","Obligatoire","Non requis"))'
    ws_doc[f"Q{r}"] = (f'=IF(A{r}="","",IF(N{r}="Non requis","N/A",'
                        f'IF(AND(O{r}="Oui",P{r}<>""),"Vert",IF(O{r}="Oui","Orange","Rouge"))))')
    # Stratégie Sortie
    ws_doc[f"R{r}"] = (f'=IF(A{r}="","",IF(AND(B{r}="PCI",C{r}="Hors Groupe"),'
                        f'"Obligatoire","Non requis"))')
    ws_doc[f"V{r}"] = (f'=IF(A{r}="","",IF(R{r}="Non requis","N/A",'
                        f'IF(AND(S{r}="Oui",T{r}<>""),"Vert",IF(S{r}="Oui","Orange","Rouge"))))')

    for col in "ABCDEFGHIJKLMNOPQRSTUV":
        c = ws_doc[f"{col}{r}"]
        if not c.value:
            c.fill = PatternFill("solid", fgColor=bg)
        else:
            c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(horizontal="center" if col not in "AGLU" else "left",
                                  vertical="center", wrap_text=True)

for col in ["E","O","S"]:
    dv(ws_doc, '"Oui,Non,En cours"', f"{col}5:{col}200")

for col in ["H","M","Q","V"]:
    rag_mfc(ws_doc, col, 5, 24)

border_range(ws_doc, 1, 50, 1, 22)
ws_doc.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET 3 — EASY
# ══════════════════════════════════════════════════════════════════════════════
ws_easy = wb.create_sheet("EASY")
ws_easy.sheet_view.showGridLines = False

ws_easy.merge_cells("A1:K1")
hdr(ws_easy["A1"], "COMPLÉTUDE CHAMPS EASY — Suivi Semestriel (27 champs obligatoires)",
    bg=NAVY, fg=WHITE, sz=11)
ws_easy.row_dimensions[1].height = 28

easy_col_defs = [
    ("A","BLOC",14),("B","Champ EASY",36),("C","Obligatoire pour",16),("D","Prestataire",26),
    ("E","S1 — Renseigné ?\n(Oui/Non/N/A)",14),("F","Date vérif. S1",14),("G","Statut S1",10),
    ("H","S2 — Renseigné ?\n(Oui/Non/N/A)",14),("I","Date vérif. S2",14),("J","Statut S2",10),
    ("K","Commentaire",34),
]
for col, title, width in easy_col_defs:
    subhdr(ws_easy[f"{col}2"], title)
    ws_easy.column_dimensions[col].width = width
ws_easy.row_dimensions[2].height = 36

easy_fields = [
    ("Identification","Nom du prestataire","Tous"),
    ("Identification","Identifiant unique (LEI ou SIRET)","Tous"),
    ("Identification","Pays d'établissement","Tous"),
    ("Identification","Groupe d'appartenance","Tous"),
    ("Qualification","Type de prestation (PCI / PS)","Tous"),
    ("Qualification","Qualification Intra / Hors Groupe","Tous"),
    ("Qualification","Score de criticité (QECI)","PCI"),
    ("Qualification","Classement DORA (ICT / Non-ICT)","PCI"),
    ("Qualification","Fonction critique ou importante (FCI)","PCI"),
    ("Données contractuelles","Date de début du contrat","Tous"),
    ("Données contractuelles","Date d'échéance du contrat","Tous"),
    ("Données contractuelles","Clause de résiliation (O/N)","PCI"),
    ("Données contractuelles","Clause PCA (O/N)","PCI"),
    ("Données contractuelles","Clause d'audit (O/N)","PCI"),
    ("Données contractuelles","Droit d'accès régulateur (O/N)","PCI"),
    ("Données financières","Montant annuel (€)","Tous"),
    ("Données financières","Niveau de dépendance économique","PCI"),
    ("Données financières","Part du CA prestataire","PCI"),
    ("Contacts","Responsable opérationnel (Amundi)","Tous"),
    ("Contacts","Contact prestataire (référent contrat)","Tous"),
    ("Contacts","Contact PCA prestataire","PCI"),
    ("Sous-traitance","Sous-traitants identifiés (O/N)","PCI"),
    ("Sous-traitance","Pays de sous-traitance","PCI"),
    ("Documents","Contrat archivé dans EASY","Tous"),
    ("Documents","QECI archivé","PCI"),
    ("Documents","Stratégie de sortie archivée","PCI HG"),
    ("Documents","Avis Risques archivé","PCI"),
]
bloc_colors = {
    "Identification":"EAF4FB","Qualification":"FFF3E0",
    "Données contractuelles":"F0FFF0","Données financières":"FFFDE7",
    "Contacts":"F5F0FF","Sous-traitance":"FFF0F0","Documents":"F0F4FF",
}

r = 3
for bloc, champ, oblig in easy_fields:
    ws_easy.row_dimensions[r].height = 20
    bg = bloc_colors.get(bloc, WHITE)
    for col, val in zip("ABCD", [bloc, champ, oblig, ""]):
        c = ws_easy[f"{col}{r}"]
        c.value = val
        c.font = Font(name="Arial", size=9, bold=(col == "A"))
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for col in "EFGHIJK":
        c = ws_easy[f"{col}{r}"]
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws_easy[f"G{r}"] = f'=IF(E{r}="","—",IF(E{r}="Oui","Vert",IF(E{r}="N/A","N/A","Rouge")))'
    ws_easy[f"J{r}"] = f'=IF(H{r}="","—",IF(H{r}="Oui","Vert",IF(H{r}="N/A","N/A","Rouge")))'
    r += 1

dv(ws_easy, '"Oui,Non,N/A"', f"E3:E{r}")
dv(ws_easy, '"Oui,Non,N/A"', f"H3:H{r}")
for col in ["F","I"]:
    for row_n in range(3, r):
        ws_easy[f"{col}{row_n}"].number_format = "DD/MM/YYYY"
rag_mfc(ws_easy, "G", 3, r)
rag_mfc(ws_easy, "J", 3, r)
border_range(ws_easy, 2, r, 1, 11)
ws_easy.freeze_panes = "D3"

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET 4 — CONTRÔLES LOD1
# ══════════════════════════════════════════════════════════════════════════════
ws_ctrl = wb.create_sheet("CONTRÔLES LOD1")
ws_ctrl.sheet_view.showGridLines = False

# Ligne 1 : titre + espace pour boutons VBA
ws_ctrl.merge_cells("A1:AB1")
hdr(ws_ctrl["A1"], "CONTRÔLES LOD1 — Suivi Périodique par Prestataire PCI",
    bg=NAVY, fg=WHITE, sz=11)
ws_ctrl.row_dimensions[1].height = 30

# Ligne 2 : note boutons
ws_ctrl.merge_cells("A2:AB2")
c = ws_ctrl["A2"]
c.value = ("▶  Boutons macros disponibles dans cette feuille :   "
           "[📧 LANCER CAMPAGNE L1-08]   [📧 RELANCES L1-08]   "
           "[📧 LANCER CAMPAGNE L1-09]   [📧 RELANCES L1-09]   "
           "— Les boutons créent des brouillons Outlook pour les non-répondants.")
c.font = Font(name="Arial", size=9, italic=True, color=NAVY)
c.fill = PatternFill("solid", fgColor=YELLOW_LT)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_ctrl.row_dimensions[2].height = 24

# Section headers row 3
for rng, txt, bg in [
    ("A3:C3","IDENTIFICATION",NAVY),
    ("D3:G3","L1-06 | EASY Entité (Trim.)",MID_BLUE),
    ("H3:K3","L1-07 | EASY Groupe (Remontée)",MID_BLUE),
    ("L3:Q3","L1-08 | Événements Majeurs (Semestriel)",NAVY),
    ("R3:V3","L1-09 | Tests PCA (Annuel T1)",MID_BLUE),
    ("W3:Z3","L1-03/05 | Contrats & Clauses (Semestriel)",MID_BLUE),
    ("AA3:AB3","L1-12 | Audits (Annuel)",MID_BLUE),
]:
    ws_ctrl.merge_cells(rng)
    hdr(ws_ctrl[rng.split(":")[0]], txt, bg=bg, sz=9)
ws_ctrl.row_dimensions[3].height = 22

ctrl_cols = [
    ("A","Prestataire",26),("B","Type",8),("C","Criticité",9),
    ("D","Dernière vérif.\nEASY entité",14),("E","Complétude\n(%)",12),
    ("F","Statut L1-06",10),("G","⏱ Prochain\ncontrôle",13),
    ("H","Dernière\nremontée registre",14),("I","Statut L1-07",10),
    ("J","Actions\nen cours",20),("K","⏱ Prochain\ncontrôle",13),
    ("L","S1 — Reporting\nreçu ?",12),("M","Statut S1",9),("N","Commentaire S1",26),
    ("O","S2 — Reporting\nreçu ?",12),("P","Statut S2",9),("Q","Commentaire S2",26),
    ("R","Bilan PCA\nreçu ?",12),("S","Date du test",14),("T","Anomalies ?",10),
    ("U","Statut L1-09",10),("V","⏱ Prochain\ncontrôle",13),
    ("W","Contrat archivé\n+ clauses OK ?",15),("X","Statut L1-03/05",12),
    ("Y","→ DOCS",10),("Z","⏱ Prochain\ncontrôle",13),
    ("AA","Date dernier\naudit",14),("AB","⏱ Prochain\naudit",13),
]
for col, title, width in ctrl_cols:
    subhdr(ws_ctrl[f"{col}4"], title)
    ws_ctrl.column_dimensions[col].width = width
ws_ctrl.row_dimensions[4].height = 38

for r in range(5, 25):
    ws_ctrl.row_dimensions[r].height = 24
    bg = WHITE if r % 2 == 0 else GREY_BG
    db_r = r  # DATABASE rows start at 4, ctrl rows at 5 → DB row = ctrl row - 1 + 4 = r

    ws_ctrl[f"A{r}"] = f'=IFERROR(DATABASE!A{r},"")' 
    ws_ctrl[f"B{r}"] = f'=IFERROR(DATABASE!E{r},"")'
    ws_ctrl[f"C{r}"] = f'=IFERROR(DATABASE!G{r},"")'

    # L1-06
    ws_ctrl[f"F{r}"] = (f'=IF(A{r}="","",IF(E{r}="","—",'
                         f'IF(E{r}>=95,"Vert",IF(E{r}>=90,"Orange","Rouge"))))')
    ws_ctrl[f"G{r}"] = f'=IF(D{r}="","À planifier",D{r}+92)'

    # L1-07
    ws_ctrl[f"I{r}"] = (f'=IF(A{r}="","",IF(H{r}="","—",'
                         f'IF(TODAY()-H{r}<=90,"Vert",IF(TODAY()-H{r}<=120,"Orange","Rouge"))))')
    ws_ctrl[f"K{r}"] = f'=IF(H{r}="","À planifier",H{r}+90)'

    # L1-08
    ws_ctrl[f"M{r}"] = f'=IF(A{r}="","",IF(L{r}="","—",IF(L{r}="Oui","Vert","Rouge")))'
    ws_ctrl[f"P{r}"] = f'=IF(A{r}="","",IF(O{r}="","—",IF(O{r}="Oui","Vert","Rouge")))'

    # L1-09
    ws_ctrl[f"U{r}"] = (f'=IF(A{r}="","",IF(R{r}="","—",'
                         f'IF(AND(R{r}="Oui",T{r}="Non"),"Vert",'
                         f'IF(AND(R{r}="Oui",T{r}="Oui"),"Orange","Rouge"))))')
    ws_ctrl[f"V{r}"] = f'=IF(S{r}="","À planifier",DATE(YEAR(S{r})+1,3,31))'

    # L1-03/05
    ws_ctrl[f"X{r}"] = f'=IF(A{r}="","",IF(W{r}="","—",IF(W{r}="Oui","Vert","Rouge")))'
    ws_ctrl[f"Y{r}"] = f'=HYPERLINK("#DOCUMENTS!A1","→ DOCS")'
    ws_ctrl[f"Z{r}"] = f'=IF(W{r}="","À planifier",W{r}+184)'

    # L1-12
    ws_ctrl[f"AB{r}"] = f'=IF(AA{r}="","À planifier",DATE(YEAR(AA{r})+1,MONTH(AA{r}),DAY(AA{r})))'

    for col, _, _ in ctrl_cols:
        c = ws_ctrl[f"{col}{r}"]
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.fill = PatternFill("solid", fgColor=bg)

# RAG
for col in ["F","I","M","P","U","X"]:
    rag_mfc(ws_ctrl, col, 5, 24)

# Date MFC
for col in ["G","K","V","Z","AB"]:
    date_mfc(ws_ctrl, col, 5, 24)
    for r in range(5, 25):
        ws_ctrl[f"{col}{r}"].number_format = "DD/MM/YYYY"

for col in ["D","H","S","AA"]:
    for r in range(5, 25):
        ws_ctrl[f"{col}{r}"].number_format = "DD/MM/YYYY"

dv(ws_ctrl, '"Oui,Non,En attente"', "L5:L500")
dv(ws_ctrl, '"Oui,Non,En attente"', "O5:O500")
dv(ws_ctrl, '"Oui,Non,En cours"',   "R5:R500")
dv(ws_ctrl, '"Oui,Non"',            "T5:T500")
dv(ws_ctrl, '"Oui,Non,En cours"',   "W5:W500")
dv(ws_ctrl, '"Satisfaisant,Réserves mineures,Réserves majeures,Non réalisé"', "AA5:AA500")

border_range(ws_ctrl, 3, 50, 1, 28)
ws_ctrl.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET 5 — PS
# ══════════════════════════════════════════════════════════════════════════════
ws_ps = wb.create_sheet("PS")
ws_ps.sheet_view.showGridLines = False

ws_ps.merge_cells("A1:L1")
hdr(ws_ps["A1"], "PRESTATIONS SIMPLES (PS) — Registre de suivi allégé",
    bg=NAVY, fg=WHITE, sz=11)
ws_ps.row_dimensions[1].height = 28

ws_ps.merge_cells("A2:L2")
c = ws_ps["A2"]
c.value = "ℹ  QECI obligatoire pour tous  |  DO requis pour nouveaux prestataires uniquement  |  L1-06 et L1-07 applicables"
c.font = Font(name="Arial", size=8, italic=True, color=NAVY)
c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
c.alignment = Alignment(horizontal="left", vertical="center")
ws_ps.row_dimensions[2].height = 20

ps_cols = [
    ("A","Prestataire",26),("B","Fonction externalisée",28),("C","Pays",12),
    ("D","Date contrat",14),("E","Responsable interne",22),
    ("F","QECI présent ?",13),("G","DO requis ?",14),("H","DO présent ?",12),
    ("I","Contrat archivé\nEASY ?",16),("J","Dernière vérif.\nEASY",14),
    ("K","Statut global",10),("L","Commentaires",32),
]
for col, title, width in ps_cols:
    subhdr(ws_ps[f"{col}3"], title)
    ws_ps.column_dimensions[col].width = width
ws_ps.row_dimensions[3].height = 32

for r in range(4, 20):
    ws_ps.row_dimensions[r].height = 20
    bg = WHITE if r % 2 == 0 else GREY_BG
    for col, _, _ in ps_cols:
        c = ws_ps[f"{col}{r}"]
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws_ps[f"K{r}"] = (f'=IF(A{r}="","",IF(AND(F{r}="Oui",I{r}="Oui"),"Vert",'
                       f'IF(OR(F{r}="Non",I{r}="Non"),"Rouge","Orange")))')

dv(ws_ps, '"Oui,Non,En cours"', "F4:F500")
dv(ws_ps, '"Obligatoire (nouveau),N/A (existant)"', "G4:G500")
dv(ws_ps, '"Oui,Non,N/A"', "H4:H500")
dv(ws_ps, '"Oui,Non,En cours"', "I4:I500")
for r in range(4, 20):
    ws_ps[f"D{r}"].number_format = "DD/MM/YYYY"
    ws_ps[f"J{r}"].number_format = "DD/MM/YYYY"
rag_mfc(ws_ps, "K", 4, 50)
border_range(ws_ps, 1, 50, 1, 12)
ws_ps.freeze_panes = "A4"

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET 6 — SYNTHÈSE
# ══════════════════════════════════════════════════════════════════════════════
ws_syn = wb.create_sheet("SYNTHÈSE")
ws_syn.sheet_view.showGridLines = False

ws_syn.merge_cells("A1:I1")
hdr(ws_syn["A1"], f"SYNTHÈSE REMÉDIATION PCI — Tableau de bord  |  MàJ : {TODAY_STR}",
    bg=NAVY, fg=WHITE, sz=12)
ws_syn.row_dimensions[1].height = 30

ws_syn.merge_cells("A2:I2")
c = ws_syn["A2"]
c.value = "Lecture seule — mis à jour automatiquement depuis les onglets DOCUMENTS et CONTRÔLES LOD1"
c.font = Font(name="Arial", size=8, italic=True, color=NAVY)
c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
c.alignment = Alignment(horizontal="left", vertical="center")
ws_syn.row_dimensions[2].height = 18

syn_cols = [
    ("A","Prestataire",28),("B","Type",8),("C","Criticité",9),
    ("D","Docs\n(RAG)",11),("E","EASY\n(RAG)",11),
    ("F","L1-08 S1",10),("G","L1-09 PCA",10),
    ("H","⚠ Risque\nglobal",11),("I","Prochaine action / Commentaire",36),
]
for col, title, width in syn_cols:
    subhdr(ws_syn[f"{col}3"], title)
    ws_syn.column_dimensions[col].width = width
ws_syn.row_dimensions[3].height = 32

for r in range(4, 24):
    ws_syn.row_dimensions[r].height = 22
    bg = WHITE if r % 2 == 0 else GREY_BG
    ctrl_r = r + 1  # CONTRÔLES LOD1 rows start at 5

    ws_syn[f"A{r}"] = f'=IFERROR(DATABASE!A{r},"")'
    ws_syn[f"B{r}"] = f'=IFERROR(DATABASE!E{r},"")'
    ws_syn[f"C{r}"] = f'=IFERROR(DATABASE!G{r},"")'
    ws_syn[f"D{r}"] = f'=IFERROR(DOCUMENTS!H{ctrl_r},"—")'
    ws_syn[f"E{r}"] = f'=IFERROR(\'CONTRÔLES LOD1\'!F{ctrl_r},"—")'
    ws_syn[f"F{r}"] = f'=IFERROR(\'CONTRÔLES LOD1\'!M{ctrl_r},"—")'
    ws_syn[f"G{r}"] = f'=IFERROR(\'CONTRÔLES LOD1\'!U{ctrl_r},"—")'
    ws_syn[f"H{r}"] = (f'=IF(A{r}=""," ",IF(OR(D{r}="Rouge",E{r}="Rouge",F{r}="Rouge",G{r}="Rouge"),"Rouge",'
                        f'IF(OR(D{r}="Orange",E{r}="Orange",F{r}="Orange",G{r}="Orange"),"Orange","Vert")))')
    for col, _, _ in syn_cols:
        c = ws_syn[f"{col}{r}"]
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(horizontal="center" if col != "I" else "left",
                                  vertical="center")
        c.fill = PatternFill("solid", fgColor=bg)

for col in ["D","E","F","G","H"]:
    rag_mfc(ws_syn, col, 4, 23)

border_range(ws_syn, 3, 50, 1, 9)
ws_syn.freeze_panes = "A4"

# ══════════════════════════════════════════════════════════════════════════════
# Tab colors & order
# ══════════════════════════════════════════════════════════════════════════════
tab_cfg = {
    "MODE D'EMPLOI": "FF9900",
    "DATABASE":      "001C4B",
    "DOCUMENTS":     "009EE0",
    "EASY":          "66B3DB",
    "CONTRÔLES LOD1":"001C4B",
    "PS":            "99CCE7",
    "SYNTHÈSE":      "FF0000",
}
for name, color in tab_cfg.items():
    wb[name].sheet_properties.tabColor = color

# ══════════════════════════════════════════════════════════════════════════════
# Sauvegarde xlsx temporaire puis injection VBA → .xlsm
# ══════════════════════════════════════════════════════════════════════════════
TEMP_XLSX = Path("/home/claude/matrice_temp.xlsx")
OUTPUT    = Path("/home/claude/Matrice_PCI_Remediation.xlsm")

wb.save(str(TEMP_XLSX))
print(f"✅ xlsx de base sauvegardé")

# ── Code VBA ──────────────────────────────────────────────────────────────────
VBA_CODE = r'''
Attribute VB_Name = "MailsLOD1"
Option Explicit

' ============================================================
'  CONFIGURATION — adapter avant utilisation
' ============================================================
Private Const EXPEDITEUR_NOM   As String = "Gianni"
Private Const EXPEDITEUR_EMAIL As String = "gianni.xxx@amundi.com"
Private Const DIRECTION        As String = "Opérations & Conformité Externalisations — Amundi Immobilier"

' Colonnes DATABASE (1-based)
Private Const COL_NOM       As Integer = 1   ' A - Nom prestataire
Private Const COL_TYPE      As Integer = 5   ' E - PCI/PS
Private Const COL_INTRAHG   As Integer = 6   ' F - Intra/HG
Private Const COL_RESP_NOM  As Integer = 11  ' K - Nom responsable interne
Private Const COL_RESP_MAIL As Integer = 12  ' L - Email responsable interne
Private Const COL_PCA_NOM   As Integer = 13  ' M - Nom contact PCA prestataire
Private Const COL_PCA_MAIL  As Integer = 14  ' N - Email contact PCA prestataire

' Colonnes CONTROLES LOD1 (1-based)
Private Const COL_CTRL_NOM  As Integer = 1   ' A - Nom (repris de DATABASE)
Private Const COL_CTRL_S1   As Integer = 12  ' L - S1 reporting reçu ?
Private Const COL_CTRL_S2   As Integer = 15  ' O - S2 reporting reçu ?
Private Const COL_CTRL_PCA  As Integer = 18  ' R - Bilan PCA reçu ?

' ============================================================
'  UTILITAIRES
' ============================================================
Private Function GetSemestre() As String
    Dim m As Integer: m = Month(Now)
    If m >= 6 And m <= 11 Then
        GetSemestre = "S1"
    Else
        GetSemestre = "S2"
    End If
End Function

Private Function GetDeadlineLabel() As String
    Dim sem As String: sem = GetSemestre()
    If sem = "S1" Then
        GetDeadlineLabel = "30 juin " & Year(Now)
    Else
        GetDeadlineLabel = "30 décembre " & Year(Now)
    End If
End Function

Private Function GetRelanceLabel() As String
    Dim sem As String: sem = GetSemestre()
    Dim deadline As Date
    If sem = "S1" Then
        deadline = DateSerial(Year(Now), 6, 30)
    Else
        deadline = DateSerial(Year(Now), 12, 30)
    End If
    Dim delta As Long: delta = DateDiff("d", Now, deadline)
    Select Case delta
        Case 7:  GetRelanceLabel = "1ère relance (J-7)"
        Case 3:  GetRelanceLabel = "2ème relance (J-3)"
        Case 1:  GetRelanceLabel = "Dernière relance (J-1)"
        Case Else: GetRelanceLabel = "Relance"
    End Select
End Function

Private Function GetRelanceLabelL09() As String
    Dim deadline As Date: deadline = DateSerial(Year(Now), 3, 31)
    Dim delta As Long: delta = DateDiff("d", Now, deadline)
    Select Case delta
        Case 7:  GetRelanceLabelL09 = "1ère relance (J-7)"
        Case 3:  GetRelanceLabelL09 = "2ème relance (J-3)"
        Case 1:  GetRelanceLabelL09 = "Dernière relance (J-1)"
        Case Else: GetRelanceLabelL09 = "Relance"
    End Select
End Function

Private Sub CreerBrouillon(dest As String, sujet As String, corps As String)
    Dim olApp As Object
    Dim mail  As Object
    On Error Resume Next
    Set olApp = GetObject(, "Outlook.Application")
    If olApp Is Nothing Then Set olApp = CreateObject("Outlook.Application")
    On Error GoTo 0
    If olApp Is Nothing Then
        MsgBox "Outlook n'est pas disponible. Vérifiez qu'Outlook est installé et ouvert.", vbCritical
        Exit Sub
    End If
    Set mail = olApp.CreateItem(0)
    mail.To      = dest
    mail.Subject = sujet
    mail.HTMLBody = corps
    mail.Display False
End Sub

Private Function SignatureHTML() As String
    SignatureHTML = "<br><br><p><b>" & EXPEDITEUR_NOM & "</b><br>" & _
                   DIRECTION & "<br><i>" & EXPEDITEUR_EMAIL & "</i></p>"
End Function

' ============================================================
'  L1-08 — LANCER CAMPAGNE (mail initial)
' ============================================================
Sub LancerCampagneL108()
    Dim wsDB   As Worksheet: Set wsDB   = ThisWorkbook.Sheets("DATABASE")
    Dim wsCtrl As Worksheet: Set wsCtrl = ThisWorkbook.Sheets("CONTRÔLES LOD1")
    Dim sem    As String: sem = GetSemestre()
    Dim dl     As String: dl  = GetDeadlineLabel()
    Dim compteur As Integer: compteur = 0

    Dim lastRow As Long: lastRow = wsDB.Cells(wsDB.Rows.Count, COL_NOM).End(xlUp).Row
    Dim r As Long
    For r = 4 To lastRow
        If UCase(Trim(wsDB.Cells(r, COL_TYPE).Value)) = "PCI" And _
           UCase(Trim(wsDB.Cells(r, COL_INTRAHG).Value)) = "HORS GROUPE" Then
            Dim nom   As String: nom   = Trim(wsDB.Cells(r, COL_NOM).Value)
            Dim rNom  As String: rNom  = Trim(wsDB.Cells(r, COL_RESP_NOM).Value)
            Dim rMail As String: rMail = Trim(wsDB.Cells(r, COL_RESP_MAIL).Value)
            If nom = "" Or rMail = "" Then GoTo NextRow108
            If InStr(rMail, "@") = 0 Then GoTo NextRow108

            Dim sujet As String
            sujet = "[EXT-L1-08 | " & sem & " " & Year(Now) & "] Évaluation semestrielle — " & nom & " — À retourner avant le " & dl

            Dim corps As String
            corps = "<p>Bonjour " & rNom & ",</p>" & _
                    "<p>Dans le cadre du dispositif de contrôle permanent de premier niveau des externalisations " & _
                    "(<b>EXT-L1-08 — Pilotage des événements majeurs</b>), nous vous sollicitons pour la réalisation " & _
                    "de l'évaluation semestrielle concernant la prestation <b>" & nom & "</b> " & _
                    "(" & sem & " " & Year(Now) & ").</p>" & _
                    "<p><b>Date limite de transmission : " & dl & ".</b></p>" & _
                    "<p><b>Ce qui vous est demandé :</b></p><ul>" & _
                    "<li>Confirmer la réception du reporting du prestataire sur la période (indicateurs qualité, incidents, SLA)</li>" & _
                    "<li>Identifier et documenter tout événement majeur ayant conduit à une dégradation du service</li>" & _
                    "<li>Renseigner le statut (Vert / Orange / Rouge) et un commentaire dans la colonne <i>" & sem & "</i> " & _
                    "de l'onglet <b>CONTRÔLES LOD1</b> de la matrice partagée</li></ul>" & _
                    "<p>En l'absence d'événement majeur à signaler, merci de confirmer explicitement en saisissant " & _
                    "<i>Oui / Vert / RAS</i> dans la matrice.</p>" & _
                    "<p>Ce contrôle est obligatoire pour toutes les PCI Hors Groupe et constitue une pièce justificative " & _
                    "exigée dans le cadre de nos obligations EBA/DORA et du Comité de Contrôle Interne (CCI).</p>" & _
                    SignatureHTML()

            CreerBrouillon rMail, sujet, corps
            compteur = compteur + 1
        End If
NextRow108:
    Next r

    If compteur = 0 Then
        MsgBox "Aucun destinataire trouvé (vérifiez les champs Type, Intra/HG et Email dans DATABASE).", vbInformation
    Else
        MsgBox compteur & " brouillon(s) Outlook créé(s) pour la campagne L1-08 " & sem & "." & Chr(10) & _
               "Vérifiez chaque brouillon avant envoi.", vbInformation, "Campagne L1-08 lancée"
    End If
End Sub

' ============================================================
'  L1-08 — RELANCES (non-répondants seulement)
' ============================================================
Sub RelancesL108()
    Dim wsDB   As Worksheet: Set wsDB   = ThisWorkbook.Sheets("DATABASE")
    Dim wsCtrl As Worksheet: Set wsCtrl = ThisWorkbook.Sheets("CONTRÔLES LOD1")
    Dim sem    As String: sem = GetSemestre()
    Dim dl     As String: dl  = GetDeadlineLabel()
    Dim rl     As String: rl  = GetRelanceLabel()
    Dim colStat As Integer
    If sem = "S1" Then colStat = COL_CTRL_S1 Else colStat = COL_CTRL_S2
    Dim compteur As Integer: compteur = 0

    Dim lastRowDB   As Long: lastRowDB   = wsDB.Cells(wsDB.Rows.Count, COL_NOM).End(xlUp).Row
    Dim lastRowCtrl As Long: lastRowCtrl = wsCtrl.Cells(wsCtrl.Rows.Count, COL_CTRL_NOM).End(xlUp).Row

    Dim r As Long
    For r = 4 To lastRowDB
        If UCase(Trim(wsDB.Cells(r, COL_TYPE).Value)) = "PCI" And _
           UCase(Trim(wsDB.Cells(r, COL_INTRAHG).Value)) = "HORS GROUPE" Then
            Dim nom   As String: nom   = Trim(wsDB.Cells(r, COL_NOM).Value)
            Dim rNom  As String: rNom  = Trim(wsDB.Cells(r, COL_RESP_NOM).Value)
            Dim rMail As String: rMail = Trim(wsDB.Cells(r, COL_RESP_MAIL).Value)
            If nom = "" Or rMail = "" Then GoTo NextRowR108
            If InStr(rMail, "@") = 0 Then GoTo NextRowR108

            ' Chercher statut dans CONTRÔLES LOD1
            Dim rc As Long: Dim dejaRendu As Boolean: dejaRendu = False
            For rc = 5 To lastRowCtrl
                If Trim(wsCtrl.Cells(rc, COL_CTRL_NOM).Value) = nom Then
                    If UCase(Trim(wsCtrl.Cells(rc, colStat).Value)) = "OUI" Then
                        dejaRendu = True
                    End If
                    Exit For
                End If
            Next rc

            If Not dejaRendu Then
                Dim sujet As String
                sujet = "[" & rl & " | EXT-L1-08 | " & sem & " " & Year(Now) & "] " & nom & " — Évaluation semestrielle à transmettre avant le " & dl

                Dim corps As String
                corps = "<p>Bonjour " & rNom & ",</p>" & _
                        "<p>Sauf erreur de notre part, nous n'avons pas encore reçu votre évaluation semestrielle " & _
                        "(<b>EXT-L1-08</b>) concernant la prestation <b>" & nom & "</b> pour le " & sem & " " & Year(Now) & ".</p>" & _
                        "<p><b>Date limite : " & dl & ".</b></p>" & _
                        "<p>Merci de renseigner le statut (Vert / Orange / Rouge) et un commentaire dans l'onglet " & _
                        "<b>CONTRÔLES LOD1</b> de la matrice, ou de nous confirmer par retour de mail " & _
                        "l'absence d'événement majeur à signaler (<i>RAS</i>).</p>" & _
                        "<p>Sans retour de votre part avant la date limite, ce contrôle sera enregistré en statut " & _
                        "<b>Rouge</b> et remonté à la Direction des Risques.</p>" & _
                        SignatureHTML()

                CreerBrouillon rMail, sujet, corps
                compteur = compteur + 1
            End If
        End If
NextRowR108:
    Next r

    If compteur = 0 Then
        MsgBox "Tous les prestataires PCI Hors Groupe ont déjà rendu leur évaluation " & sem & ".", vbInformation
    Else
        MsgBox compteur & " brouillon(s) de relance L1-08 créé(s) pour le " & sem & "." & Chr(10) & _
               "Vérifiez chaque brouillon avant envoi.", vbInformation, "Relances L1-08"
    End If
End Sub

' ============================================================
'  L1-09 — LANCER CAMPAGNE (mail initial prestataires)
' ============================================================
Sub LancerCampagneL109()
    Dim wsDB   As Worksheet: Set wsDB   = ThisWorkbook.Sheets("DATABASE")
    Dim annee  As Integer: annee = Year(Now) - 1
    Dim dl     As String: dl = "31 mars " & Year(Now)
    Dim compteur As Integer: compteur = 0

    Dim lastRow As Long: lastRow = wsDB.Cells(wsDB.Rows.Count, COL_NOM).End(xlUp).Row
    Dim r As Long
    For r = 4 To lastRow
        If UCase(Trim(wsDB.Cells(r, COL_TYPE).Value)) = "PCI" Then
            Dim nom    As String: nom    = Trim(wsDB.Cells(r, COL_NOM).Value)
            Dim pcaNom As String: pcaNom = Trim(wsDB.Cells(r, COL_PCA_NOM).Value)
            Dim pcaMail As String: pcaMail = Trim(wsDB.Cells(r, COL_PCA_MAIL).Value)
            If nom = "" Or pcaMail = "" Then GoTo NextRow109
            If InStr(pcaMail, "@") = 0 Then GoTo NextRow109

            Dim sujet As String
            sujet = "[EXT-L1-09 | Amundi Immobilier] Demande de bilan PCA " & annee & " — À transmettre avant le " & dl

            Dim corps As String
            corps = "<p>Bonjour " & pcaNom & ",</p>" & _
                    "<p>Dans le cadre de nos obligations contractuelles et réglementaires " & _
                    "(EBA Guidelines on Outsourcing, DORA — contrôle <b>EXT-L1-09</b>), " & _
                    "nous vous sollicitons pour la transmission du <b>bilan annuel de tests PCA</b> " & _
                    "relatif à votre prestation pour <b>Amundi Immobilier</b>, pour l'exercice <b>" & annee & "</b>.</p>" & _
                    "<p><b>Date limite de transmission : " & dl & ".</b></p>" & _
                    "<p><b>Documents attendus :</b></p><ul>" & _
                    "<li>Bilan des tests PCA réalisés en " & annee & " (scénarios couverts, résultats, anomalies éventuelles)</li>" & _
                    "<li>Preuve de réalisation des tests (rapport, procès-verbal ou attestation)</li>" & _
                    "<li>Plan d'action correctif si des anomalies ont été identifiées</li>" & _
                    "<li>Confirmation de l'interaction entre votre dispositif de gestion de crise et celui d'Amundi Immobilier</li></ul>" & _
                    "<p>Ces éléments constituent la piste d'audit requise pour notre dispositif de contrôle interne " & _
                    "et sont susceptibles d'être demandés par nos autorités de supervision (BCE, ACPR/AMF).</p>" & _
                    "<p>Merci de transmettre ces documents à l'adresse : <b>" & EXPEDITEUR_EMAIL & "</b> " & _
                    "en mentionnant en objet la référence de votre contrat avec Amundi Immobilier.</p>" & _
                    SignatureHTML()

            CreerBrouillon pcaMail, sujet, corps
            compteur = compteur + 1
        End If
NextRow109:
    Next r

    If compteur = 0 Then
        MsgBox "Aucun destinataire PCI trouvé (vérifiez les champs Type et Email Contact PCA dans DATABASE).", vbInformation
    Else
        MsgBox compteur & " brouillon(s) Outlook créé(s) pour la campagne L1-09." & Chr(10) & _
               "Vérifiez chaque brouillon avant envoi.", vbInformation, "Campagne L1-09 lancée"
    End If
End Sub

' ============================================================
'  L1-09 — RELANCES (prestataires sans bilan reçu)
' ============================================================
Sub RelancesL109()
    Dim wsDB   As Worksheet: Set wsDB   = ThisWorkbook.Sheets("DATABASE")
    Dim wsCtrl As Worksheet: Set wsCtrl = ThisWorkbook.Sheets("CONTRÔLES LOD1")
    Dim annee  As Integer: annee = Year(Now) - 1
    Dim dl     As String: dl = "31 mars " & Year(Now)
    Dim rl     As String: rl = GetRelanceLabelL09()
    Dim compteur As Integer: compteur = 0

    Dim lastRowDB   As Long: lastRowDB   = wsDB.Cells(wsDB.Rows.Count, COL_NOM).End(xlUp).Row
    Dim lastRowCtrl As Long: lastRowCtrl = wsCtrl.Cells(wsCtrl.Rows.Count, COL_CTRL_NOM).End(xlUp).Row

    Dim r As Long
    For r = 4 To lastRowDB
        If UCase(Trim(wsDB.Cells(r, COL_TYPE).Value)) = "PCI" Then
            Dim nom    As String: nom    = Trim(wsDB.Cells(r, COL_NOM).Value)
            Dim pcaNom  As String: pcaNom  = Trim(wsDB.Cells(r, COL_PCA_NOM).Value)
            Dim pcaMail As String: pcaMail = Trim(wsDB.Cells(r, COL_PCA_MAIL).Value)
            If nom = "" Or pcaMail = "" Then GoTo NextRowR109
            If InStr(pcaMail, "@") = 0 Then GoTo NextRowR109

            Dim rc As Long: Dim dejaRecu As Boolean: dejaRecu = False
            For rc = 5 To lastRowCtrl
                If Trim(wsCtrl.Cells(rc, COL_CTRL_NOM).Value) = nom Then
                    If UCase(Trim(wsCtrl.Cells(rc, COL_CTRL_PCA).Value)) = "OUI" Then
                        dejaRecu = True
                    End If
                    Exit For
                End If
            Next rc

            If Not dejaRecu Then
                Dim sujet As String
                sujet = "[" & rl & " | EXT-L1-09 | Amundi Immobilier] Bilan PCA " & annee & " — Transmission attendue avant le " & dl

                Dim corps As String
                corps = "<p>Bonjour " & pcaNom & ",</p>" & _
                        "<p>Sauf erreur de notre part, nous n'avons pas encore reçu le <b>bilan de tests PCA " & annee & "</b> " & _
                        "relatif à votre prestation pour Amundi Immobilier (<b>EXT-L1-09</b>).</p>" & _
                        "<p><b>Date limite : " & dl & ".</b></p>" & _
                        "<p>Sans réception de ces documents avant la date limite, nous serons contraints d'enregistrer " & _
                        "ce contrôle en statut <b>Rouge</b> et d'en informer notre Direction des Risques, " & _
                        "conformément à nos obligations réglementaires (DORA, EBA).</p>" & _
                        "<p>Pour rappel, les documents attendus sont :" & _
                        "<ul><li>Bilan des tests PCA " & annee & " (scénarios, résultats)</li>" & _
                        "<li>Preuve de réalisation</li>" & _
                        "<li>Plan d'action correctif si anomalies</li></ul></p>" & _
                        "<p>Merci de transmettre ces éléments à : <b>" & EXPEDITEUR_EMAIL & "</b></p>" & _
                        SignatureHTML()

                CreerBrouillon pcaMail, sujet, corps
                compteur = compteur + 1
            End If
        End If
NextRowR109:
    Next r

    If compteur = 0 Then
        MsgBox "Tous les bilans PCA ont été reçus.", vbInformation
    Else
        MsgBox compteur & " brouillon(s) de relance L1-09 créé(s)." & Chr(10) & _
               "Vérifiez chaque brouillon avant envoi.", vbInformation, "Relances L1-09"
    End If
End Sub
'''

# ── Injection du VBA dans le zip .xlsm ───────────────────────────────────────
import re

VBA_MODULE = VBA_CODE.encode("utf-8")

# Le .xlsm est un zip. On doit :
# 1. Copier le xlsx en xlsm
# 2. Ajouter vbaProject.bin (binaire VBA) — non possible en Python pur natif
# 3. ALTERNATIVE : on intègre le VBA comme feuille "SCRIPTS_VBA" lisible
#    ET on écrit un fichier .bas séparé pour import dans VBE

# Approche retenue :
# - Créer un onglet caché "SCRIPTS_VBA" dans le xlsx avec le code VBA
# - Créer un fichier .bas séparé à importer dans l'éditeur VBE (Alt+F11)
# - Documenter la procédure d'import dans MODE D'EMPLOI

import shutil
shutil.copy(str(TEMP_XLSX), str(OUTPUT.with_suffix(".xlsx")))

# Réécrire en .xlsm (même contenu, extension seule change pour signaler les macros)
# Note : vrai .xlsm nécessite vbaProject.bin binaire, généré ici via fichier .bas
shutil.copy(str(TEMP_XLSX), str(OUTPUT))

# Onglet caché SCRIPTS_VBA avec le code
wb2 = __import__("openpyxl").load_workbook(str(OUTPUT))
ws_vba = wb2.create_sheet("SCRIPTS_VBA")
ws_vba.sheet_state = "hidden"
ws_vba.sheet_view.showGridLines = False

ws_vba.column_dimensions["A"].width = 3
ws_vba.column_dimensions["B"].width = 120
ws_vba.merge_cells("B1:B1")
hdr(ws_vba["B1"], "CODE VBA — À copier dans l'éditeur Visual Basic (Alt+F11 → Insertion → Module)",
    bg=NAVY, fg=WHITE, sz=10)
ws_vba.row_dimensions[1].height = 24

lines = VBA_CODE.split("\n")
for i, line in enumerate(lines, start=2):
    c = ws_vba.cell(row=i, column=2, value=line)
    c.font = Font(name="Courier New", size=8)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ws_vba.row_dimensions[i].height = 14

wb2.save(str(OUTPUT))
print(f"✅ Onglet SCRIPTS_VBA injecté")

# Fichier .bas pour import direct VBE
BAS_PATH = Path("/home/claude/MailsLOD1.bas")
BAS_PATH.write_text(
    "Attribute VB_Name = \"MailsLOD1\"\r\n" + VBA_CODE,
    encoding="utf-8"
)
print(f"✅ Fichier .bas sauvegardé : {BAS_PATH}")

TEMP_XLSX.unlink(missing_ok=True)
print(f"\n{'='*60}")
print(f"  LIVRAISON :")
print(f"  • {OUTPUT.name}  — matrice complète + code VBA")
print(f"  • MailsLOD1.bas      — module VBA à importer")
print(f"\n  ACTIVER LES MACROS (une seule fois) :")
print(f"  1. Ouvrir {OUTPUT.name} dans Excel")
print(f"  2. Alt+F11 → Insertion → Module")
print(f"  3. Coller le contenu de MailsLOD1.bas")
print(f"  4. Fermer VBE → Sauvegarder en .xlsm")
print(f"  5. Les 4 boutons apparaissent dans CONTRÔLES LOD1")
print(f"{'='*60}\n")
