import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule, CellIsRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation
import datetime

wb = Workbook()

# ── Palette Amundi ──────────────────────────────────────────────────────────
BLUE       = "009EE0"
NAVY       = "001C4B"
LIGHT_BLUE = "CCE5F3"
MID_BLUE   = "66B3DB"
WHITE      = "FFFFFF"
GREY_BG    = "F5F7FA"
GREY_BORDER= "BFCAD4"
GREEN      = "00B050"
ORANGE     = "FF9900"
RED        = "FF0000"
YELLOW     = "FFFF00"
RED_LIGHT  = "FFD7D7"
ORANGE_LT  = "FFE5B4"
GREEN_LT   = "D6F0DD"

def hdr(cell, txt, bg=NAVY, fg=WHITE, sz=10, bold=True, wrap=True, center=True):
    cell.value = txt
    cell.font = Font(name="Arial", bold=bold, color=fg, size=sz)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center" if center else "left",
                                vertical="center", wrap_text=wrap)

def subhdr(cell, txt, bg=LIGHT_BLUE, fg=NAVY, sz=9, bold=True):
    cell.value = txt
    cell.font = Font(name="Arial", bold=bold, color=fg, size=sz)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def data_cell(cell, txt="", sz=9, bold=False, center=False, bg=None):
    cell.value = txt
    cell.font = Font(name="Arial", size=sz, bold=bold)
    cell.alignment = Alignment(horizontal="center" if center else "left",
                                vertical="center", wrap_text=True)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)

def thin_border(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color=GREY_BORDER)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def add_dv(ws, formula, sqref, prompt="", title=""):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True,
                        showErrorMessage=True, showInputMessage=bool(prompt))
    dv.sqref = sqref
    if prompt:
        dv.promptTitle = title
        dv.prompt = prompt
    ws.add_data_validation(dv)
    return dv

def rag_mfc(ws, col_letter, row_start, row_end):
    green_fill = PatternFill("solid", fgColor=GREEN_LT)
    orange_fill= PatternFill("solid", fgColor=ORANGE_LT)
    red_fill   = PatternFill("solid", fgColor=RED_LIGHT)
    rng = f"{col_letter}{row_start}:{col_letter}{row_end}"
    ws.conditional_formatting.add(rng, CellIsRule("equal", ['"Vert"'],  fill=green_fill))
    ws.conditional_formatting.add(rng, CellIsRule("equal", ['"Orange"'],fill=orange_fill))
    ws.conditional_formatting.add(rng, CellIsRule("equal", ['"Rouge"'], fill=red_fill))

# ────────────────────────────────────────────────────────────────────────────
# 1. DATABASE
# ────────────────────────────────────────────────────────────────────────────
ws_db = wb.active
ws_db.title = "DATABASE"
ws_db.sheet_view.showGridLines = False
ws_db.row_dimensions[1].height = 14
ws_db.row_dimensions[2].height = 38
ws_db.row_dimensions[3].height = 34

# Title banner
ws_db.merge_cells("A1:U1")
hdr(ws_db["A1"], "MATRICE DE SUIVI REMÉDIATION PCI — Amundi Immobilier",
    bg=NAVY, fg=WHITE, sz=13)

# Section headers row 2
sections = [
    ("A2:D2",  "IDENTIFICATION PRESTATAIRE", NAVY),
    ("E2:G2",  "QUALIFICATION", NAVY),
    ("H2:J2",  "CONTRAT", NAVY),
    ("K2:L2",  "RESPONSABLE INTERNE", NAVY),
    ("M2:N2",  "CONTACT PCA PRESTATAIRE", NAVY),
    ("O2:P2",  "SUIVI FINANCIER", NAVY),
    ("Q2:R2",  "STATUTS", NAVY),
    ("S2:U2",  "MÉTADONNÉES", NAVY),
]
for rng, txt, bg in sections:
    ws_db.merge_cells(rng)
    hdr(ws_db[rng.split(":")[0]], txt, bg=bg, sz=9)

# Column headers row 3
db_cols = [
    ("A", "Nom Prestataire", 28),
    ("B", "Code Interne", 14),
    ("C", "Groupe / Entité mère", 20),
    ("D", "Pays", 12),
    ("E", "Type\n(PCI / PS)", 12),
    ("F", "Intra / Hors Groupe", 14),
    ("G", "Criticité\n(1→4)", 10),
    ("H", "Fonction Externalisée", 28),
    ("I", "Date Entrée en Vigueur", 16),
    ("J", "Date Échéance Contrat", 16),
    ("K", "Responsable Opérationnel\n(Nom)", 22),
    ("L", "Email Responsable\nOpérationnel", 28),
    ("M", "Contact PCA\nPrestataire (Nom)", 22),
    ("N", "Email Contact PCA\nPrestataire", 28),
    ("O", "Montant Annuel (€)", 16),
    ("P", "Dépendance Éco.", 14),
    ("Q", "Statut Prestataire\n(Actif / Inactif / En cours)", 18),
    ("R", "Statut Contrat\n(Valide / Échu / Résiliation)", 18),
    ("S", "Date Création\nFiche", 14),
    ("T", "Dernière MàJ", 14),
    ("U", "Commentaires", 30),
]
for col, title, width in db_cols:
    subhdr(ws_db[f"{col}3"], title)
    ws_db.column_dimensions[col].width = width

# Sample data rows (5 prestataires fictifs)
sample = [
    ["CORTEX / UPTEVIA","PCI-001","Groupe La Poste","France","PCI","Hors Groupe",3,
     "Impression / Routage courriers actionnaires","2024-01-01","2026-12-31",
     "Julien Manuel","julien.manuel@amundi.com","Sophie Leclerc","s.leclerc@cortex.fr",
     120000,"Élevée","Actif","Valide","2025-01-10","2025-06-01","Transition Démat en cours"],
    ["HIVER","PCI-002","Indépendant","France","PCI","Hors Groupe",4,
     "Impression / Routage (ancien prestataire)","2019-03-01","2025-12-31",
     "Julien Manuel","julien.manuel@amundi.com","Pierre Hiver","contact@hiver.fr",
     95000,"Critique","Inactif","Résiliation","2025-01-10","2025-06-01","Procédure liquidation judiciaire"],
    ["SOPRA BANKING","PCI-003","Sopra Steria Group","France","PCI","Hors Groupe",3,
     "Logiciel de gestion actionnaires (Espace Privé)","2021-06-01","2027-05-31",
     "Murielle Dugois","murielle.dugois@amundi.com","Ref. Contrat Sopra","contact@sopra.fr",
     280000,"Élevée","Actif","Valide","2025-01-10","2025-06-15","KPI en cours de définition"],
    ["KPMG","PCI-004","KPMG International","France","PCI","Hors Groupe",2,
     "Commissariat aux comptes SCPI","2022-01-01","2027-12-31",
     "Jean-Marc Fayet","jm.fayet@amundi.com","Ref. Mission KPMG","audit@kpmg.fr",
     150000,"Modérée","Actif","Valide","2025-01-10","2025-06-01","RAS"],
    ["BNP PARIBAS SECURITIES","PCI-005","BNP Paribas","France","PCI","Hors Groupe",3,
     "Conservation de titres","2020-09-01","2026-08-31",
     "Edouard Auche","edouard.auche@amundi.com","Ref. Opérations BNP","ops@bnpparibas.com",
     200000,"Élevée","Actif","Valide","2025-01-10","2025-06-01","Renouvellement à préparer 2026"],
]

for i, row in enumerate(sample, start=4):
    ws_db.row_dimensions[i].height = 22
    for j, val in enumerate(row, start=1):
        c = ws_db.cell(row=i, column=j, value=val)
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.fill = PatternFill("solid", fgColor=WHITE if i % 2 == 0 else GREY_BG)

# Dropdowns DATABASE
add_dv(ws_db, '"PCI,PS"', "E4:E200", "PCI = Critique/Importante · PS = Simple", "Type")
add_dv(ws_db, '"Hors Groupe,Intragroupe"', "F4:F200")
add_dv(ws_db, '"1,2,3,4"', "G4:G200", "1=Faible 2=Modéré 3=Élevé 4=Critique", "Criticité")
add_dv(ws_db, '"Actif,Inactif,En cours"', "Q4:Q200")
add_dv(ws_db, '"Valide,Échu,Résiliation,En renouvellement"', "R4:R200")

thin_border(ws_db, 2, 50, 1, 21)

# ────────────────────────────────────────────────────────────────────────────
# 2. DOCUMENTS
# ────────────────────────────────────────────────────────────────────────────
ws_doc = wb.create_sheet("DOCUMENTS")
ws_doc.sheet_view.showGridLines = False

ws_doc.merge_cells("A1:V1")
hdr(ws_doc["A1"], "COMPLÉTUDE DOCUMENTAIRE — PCI (Dossier d'Externalisation)",
    bg=NAVY, fg=WHITE, sz=12)

# Section headers row 2
doc_sections = [
    ("A2:C2",  "IDENTIFICATION", NAVY),
    ("D2:H2",  "DOSSIER D'OPPORTUNITÉ", MID_BLUE),
    ("I2:M2",  "QECI", MID_BLUE),
    ("N2:Q2",  "AVIS RISQUES (PCI)", MID_BLUE),
    ("R2:V2",  "STRATÉGIE DE SORTIE (PCI HG)", MID_BLUE),
]
for rng, txt, bg in doc_sections:
    ws_doc.merge_cells(rng)
    hdr(ws_doc[rng.split(":")[0]], txt, bg=bg, sz=9)

# Col headers row 3
doc_cols = [
    ("A","Prestataire",26), ("B","Type",8), ("C","Intra/HG",10),
    ("D","Requis ?",10), ("E","Présent ?",10), ("F","Date",12),
    ("G","Lien GED",22), ("H","Statut",10),
    ("I","Requis ?",10), ("J","Version",12), ("K","Date QECI",12),
    ("L","Lien GED",22), ("M","Statut",10),
    ("N","Requis ?",10), ("O","Avis rendu ?",10), ("P","Date",12),
    ("Q","Statut",10),
    ("R","Requis ?",10), ("S","Présente ?",10), ("T","Date",12),
    ("U","Lien GED",22), ("V","Statut",10),
]
for col, title, width in doc_cols:
    subhdr(ws_doc[f"{col}3"], title)
    ws_doc.column_dimensions[col].width = width
ws_doc.row_dimensions[3].height = 32

# Instructions row 4
ws_doc.merge_cells("A4:V4")
c = ws_doc["A4"]
c.value = ("ℹ  Règles d'obligation : DO = N/A pour prestataires existants, Obligatoire pour nouveaux  |  "
           "QECI = Obligatoire pour tous  |  Avis Risques = PCI uniquement  |  "
           "Stratégie de Sortie = PCI Hors Groupe uniquement")
c.font = Font(name="Arial", size=8, italic=True, color=NAVY)
c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_doc.row_dimensions[4].height = 28

# Formulas for each PCI row (rows 5-9 = 5 prestataires)
for i, name in enumerate(["CORTEX / UPTEVIA","HIVER","SOPRA BANKING","KPMG","BNP PARIBAS SECURITIES"], start=5):
    r = i
    ws_doc.row_dimensions[r].height = 22
    bg = WHITE if r % 2 == 0 else GREY_BG

    # Identification (pulls from DATABASE)
    ws_doc[f"A{r}"] = f"=DATABASE!A{r-1}"
    ws_doc[f"B{r}"] = f"=DATABASE!E{r-1}"
    ws_doc[f"C{r}"] = f"=DATABASE!F{r-1}"

    # DO — requis : N/A si prestataire existant (B=PCI et sample rows = existants → N/A)
    ws_doc[f"D{r}"] = f'=IF(B{r}="PCI",IF(DATABASE!I{r-1}<>"","N/A (existant)","Obligatoire"),"Non requis")'
    ws_doc[f"E{r}"] = ""
    ws_doc[f"F{r}"] = ""
    ws_doc[f"G{r}"] = ""
    ws_doc[f"H{r}"] = f'=IF(D{r}="N/A (existant)","N/A",IF(AND(E{r}="Oui",F{r}<>""),"Vert",IF(E{r}="Oui","Orange","Rouge")))'

    # QECI
    ws_doc[f"I{r}"] = "Obligatoire"
    ws_doc[f"J{r}"] = ""
    ws_doc[f"K{r}"] = ""
    ws_doc[f"L{r}"] = ""
    ws_doc[f"M{r}"] = f'=IF(AND(J{r}<>"",K{r}<>""),"Vert",IF(J{r}<>"","Orange","Rouge"))'

    # Avis Risques
    ws_doc[f"N{r}"] = f'=IF(B{r}="PCI","Obligatoire","Non requis")'
    ws_doc[f"O{r}"] = ""
    ws_doc[f"P{r}"] = ""
    ws_doc[f"Q{r}"] = f'=IF(N{r}="Non requis","N/A",IF(AND(O{r}="Oui",P{r}<>""),"Vert",IF(O{r}="Oui","Orange","Rouge")))'

    # Stratégie de Sortie
    ws_doc[f"R{r}"] = f'=IF(AND(B{r}="PCI",C{r}="Hors Groupe"),"Obligatoire","Non requis")'
    ws_doc[f"S{r}"] = ""
    ws_doc[f"T{r}"] = ""
    ws_doc[f"U{r}"] = ""
    ws_doc[f"V{r}"] = f'=IF(R{r}="Non requis","N/A",IF(AND(S{r}="Oui",T{r}<>""),"Vert",IF(S{r}="Oui","Orange","Rouge")))'

    for col in "ABCDEFGHIJKLMNOPQRSTUV":
        c = ws_doc[f"{col}{r}"]
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(horizontal="left" if col in "AGI" else "center",
                                  vertical="center", wrap_text=True)
        c.fill = PatternFill("solid", fgColor=bg)

# Dropdowns DOCUMENTS
for col in ["E","O","S"]:
    add_dv(ws_doc, '"Oui,Non,En cours"', f"{col}5:{col}200")

# RAG MFC
for col, row_s, row_e in [("H",5,9),("M",5,9),("Q",5,9),("V",5,9)]:
    rag_mfc(ws_doc, col, row_s, row_e)

thin_border(ws_doc, 2, 50, 1, 22)

# ────────────────────────────────────────────────────────────────────────────
# 3. EASY
# ────────────────────────────────────────────────────────────────────────────
ws_easy = wb.create_sheet("EASY")
ws_easy.sheet_view.showGridLines = False

ws_easy.merge_cells("A1:K1")
hdr(ws_easy["A1"], "COMPLÉTUDE CHAMPS EASY — Suivi Semestriel par Prestataire",
    bg=NAVY, fg=WHITE, sz=12)

easy_col_defs = [
    ("A","BLOC",14), ("B","Champ EASY",34),
    ("C","Obligatoire pour",16), ("D","Prestataire",24),
    ("E","Valeur renseignée\nS1 (O/N)",14), ("F","Date vérif. S1",14), ("G","Statut S1",10),
    ("H","Valeur renseignée\nS2 (O/N)",14), ("I","Date vérif. S2",14), ("J","Statut S2",10),
    ("K","Commentaire",32),
]
for col, title, width in easy_col_defs:
    subhdr(ws_easy[f"{col}2"], title)
    ws_easy.column_dimensions[col].width = width
ws_easy.row_dimensions[2].height = 34

# EASY champs obligatoires
easy_fields = [
    # (bloc, champ, obligation)
    ("Identification","Nom du prestataire","Tous"),
    ("Identification","Identifiant unique (LEI ou SIRET)","Tous"),
    ("Identification","Pays d'établissement","Tous"),
    ("Identification","Groupe d'appartenance","Tous"),
    ("Qualification","Type de prestation (PCI / PS)","Tous"),
    ("Qualification","Qualification Intra / Hors Groupe","Tous"),
    ("Qualification","Score de criticité (QECI)","PCI"),
    ("Qualification","Classement DORA (ICT / Non-ICT)","PCI"),
    ("Qualification","Fonction critique ou importante (FCI)","PCI"),
    ("Données contractuelles","Date de début du contrat","Tous"),
    ("Données contractuelles","Date d'échéance du contrat","Tous"),
    ("Données contractuelles","Clause de résiliation (O/N)","PCI"),
    ("Données contractuelles","Clause PCA (O/N)","PCI"),
    ("Données contractuelles","Clause d'audit (O/N)","PCI"),
    ("Données contractuelles","Droit d'accès régulateur (O/N)","PCI"),
    ("Données financières","Montant annuel (€)","Tous"),
    ("Données financières","Niveau de dépendance économique","PCI"),
    ("Données financières","Part du CA prestataire","PCI"),
    ("Contacts","Responsable opérationnel (Amundi)","Tous"),
    ("Contacts","Contact prestataire (référent contrat)","Tous"),
    ("Contacts","Contact PCA prestataire","PCI"),
    ("Sous-traitance","Sous-traitants identifiés (O/N)","PCI"),
    ("Sous-traitance","Pays de sous-traitance","PCI"),
    ("Documents","Contrat archivé dans EASY","Tous"),
    ("Documents","QECI archivé","PCI"),
    ("Documents","Stratégie de sortie archivée","PCI HG"),
    ("Documents","Avis Risques archivé","PCI"),
]

bloc_colors = {
    "Identification":       "EAF4FB",
    "Qualification":        "FFF3E0",
    "Données contractuelles":"F0FFF0",
    "Données financières":  "FFF9E6",
    "Contacts":             "F5F0FF",
    "Sous-traitance":       "FFF0F0",
    "Documents":            "F0F4FF",
}

row = 3
for bloc, champ, oblig in easy_fields:
    ws_easy.row_dimensions[row].height = 20
    bg = bloc_colors.get(bloc, WHITE)
    for col, val in zip("ABCD", [bloc, champ, oblig, ""]):
        c = ws_easy[f"{col}{row}"]
        c.value = val
        c.font = Font(name="Arial", size=9, bold=(col == "A"))
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for col in "EFGHIJ":
        c = ws_easy[f"{col}{row}"]
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Statut S1
    ws_easy[f"G{row}"] = f'=IF(E{row}="","—",IF(E{row}="Oui","Vert","Rouge"))'
    # Statut S2
    ws_easy[f"J{row}"] = f'=IF(H{row}="","—",IF(H{row}="Oui","Vert","Rouge"))'

    ws_easy[f"K{row}"].font = Font(name="Arial", size=9)
    ws_easy[f"K{row}"].fill = PatternFill("solid", fgColor=bg)
    row += 1

add_dv(ws_easy, '"Oui,Non,N/A"', f"E3:E{row}", "O = renseigné · N = manquant · N/A = non applicable", "S1")
add_dv(ws_easy, '"Oui,Non,N/A"', f"H3:H{row}")
rag_mfc(ws_easy, "G", 3, row)
rag_mfc(ws_easy, "J", 3, row)
thin_border(ws_easy, 2, row, 1, 11)

# ────────────────────────────────────────────────────────────────────────────
# 4. CONTRÔLES LOD1
# ────────────────────────────────────────────────────────────────────────────
ws_ctrl = wb.create_sheet("CONTRÔLES LOD1")
ws_ctrl.sheet_view.showGridLines = False

ws_ctrl.merge_cells("A1:AD1")
hdr(ws_ctrl["A1"], "CONTRÔLES LOD1 — Suivi Périodique par Prestataire PCI",
    bg=NAVY, fg=WHITE, sz=12)

# Section row 2
ctrl_sections = [
    ("A2:C2",  "IDENTIFICATION",             NAVY),
    ("D2:G2",  "L1-06 | EASY Entité (Trim.)",MID_BLUE),
    ("H2:K2",  "L1-07 | EASY Groupe (Remontée)",MID_BLUE),
    ("L2:Q2",  "L1-08 | Événements Majeurs (Sem.)",NAVY),
    ("R2:V2",  "L1-09 | Tests PCA (Annuel T1)",MID_BLUE),
    ("W2:Z2",  "L1-03/05 | Contrats & Clauses (Sem.)",MID_BLUE),
    ("AA2:AD2","L1-12 | Audits Fournisseurs (Annuel)",MID_BLUE),
]
for rng, txt, bg in ctrl_sections:
    ws_ctrl.merge_cells(rng)
    hdr(ws_ctrl[rng.split(":")[0]], txt, bg=bg, sz=9)

ctrl_cols = [
    ("A","Prestataire",26), ("B","Type",8), ("C","Criticité",9),
    ("D","Dernière vérif.\nEASY entité",14), ("E","Taux complétude\n(estimé %)",14),
    ("F","Statut L1-06",10), ("G","⏱ Prochain\ncontrôle",12),
    ("H","Dernière remontée\nregistre Groupe",14), ("I","Statut L1-07",10),
    ("J","Actions\nen cours",20), ("K","⏱ Prochain\ncontrôle",12),
    ("L","S1 — Reporting\nreçu ?",12), ("M","Statut S1",9), ("N","Commentaire S1",24),
    ("O","S2 — Reporting\nreçu ?",12), ("P","Statut S2",9), ("Q","Commentaire S2",24),
    ("R","Bilan PCA\nreçu ?",12), ("S","Date du test\nPCA",14), ("T","Anomalies\n?",10),
    ("U","Statut L1-09",10), ("V","⏱ Prochain\ncontrôle",12),
    ("W","Contrat archivé\n+ clauses OK ?",14), ("X","Statut L1-03/05",10),
    ("Y","Lien onglet\nDOCS",14), ("Z","⏱ Prochain\ncontrôle",12),
    ("AA","Date dernier\naudit",14), ("AB","Résultat\naudit",14),
    ("AC","Plan d'action\nsuivi ?",12), ("AD","⏱ Prochain\ncontrôle",12),
]
for col, title, width in ctrl_cols:
    subhdr(ws_ctrl[f"{col}3"], title)
    ws_ctrl.column_dimensions[col].width = width
ws_ctrl.row_dimensions[3].height = 38

for i in range(5):
    r = i + 4
    ws_ctrl.row_dimensions[r].height = 24
    bg = WHITE if r % 2 == 0 else GREY_BG

    ws_ctrl[f"A{r}"] = f"=DATABASE!A{i+4}"
    ws_ctrl[f"B{r}"] = f"=DATABASE!E{i+4}"
    ws_ctrl[f"C{r}"] = f"=DATABASE!G{i+4}"

    # L1-06 — prochain contrôle = +3 mois
    ws_ctrl[f"D{r}"] = ""
    ws_ctrl[f"E{r}"] = ""
    ws_ctrl[f"F{r}"] = f'=IF(D{r}="","—",IF(E{r}>=95,"Vert",IF(E{r}>=90,"Orange","Rouge")))'
    ws_ctrl[f"G{r}"] = f'=IF(D{r}="","À planifier",D{r}+92)'

    # L1-07
    ws_ctrl[f"H{r}"] = ""
    ws_ctrl[f"I{r}"] = f'=IF(H{r}="","—",IF(TODAY()-H{r}<=90,"Vert",IF(TODAY()-H{r}<=120,"Orange","Rouge")))'
    ws_ctrl[f"J{r}"] = ""
    ws_ctrl[f"K{r}"] = f'=IF(H{r}="","À planifier",H{r}+90)'

    # L1-08 S1/S2
    ws_ctrl[f"L{r}"] = ""
    ws_ctrl[f"M{r}"] = f'=IF(L{r}="","—",IF(L{r}="Oui","Vert","Rouge"))'
    ws_ctrl[f"N{r}"] = ""
    ws_ctrl[f"O{r}"] = ""
    ws_ctrl[f"P{r}"] = f'=IF(O{r}="","—",IF(O{r}="Oui","Vert","Rouge"))'
    ws_ctrl[f"Q{r}"] = ""

    # L1-09
    ws_ctrl[f"R{r}"] = ""
    ws_ctrl[f"S{r}"] = ""
    ws_ctrl[f"T{r}"] = ""
    ws_ctrl[f"U{r}"] = f'=IF(R{r}="","—",IF(AND(R{r}="Oui",T{r}="Non"),"Vert",IF(AND(R{r}="Oui",T{r}="Oui"),"Orange","Rouge")))'
    ws_ctrl[f"V{r}"] = f'=IF(S{r}="","À planifier",DATE(YEAR(S{r})+1,3,31))'

    # L1-03/05
    ws_ctrl[f"W{r}"] = ""
    ws_ctrl[f"X{r}"] = f'=IF(W{r}="","—",IF(W{r}="Oui","Vert","Rouge"))'
    ws_ctrl[f"Y{r}"] = f'=HYPERLINK("DOCUMENTS","→ DOCS")'
    ws_ctrl[f"Z{r}"] = f'=IF(W{r}="","À planifier",W{r}+184)'

    # L1-12
    ws_ctrl[f"AA{r}"] = ""
    ws_ctrl[f"AB{r}"] = ""
    ws_ctrl[f"AC{r}"] = ""
    ws_ctrl[f"AD{r}"] = f'=IF(AA{r}="","À planifier",DATE(YEAR(AA{r})+1,MONTH(AA{r}),DAY(AA{r})))'

    for col_letter in [c[0] for c in ctrl_cols]:
        c = ws_ctrl[f"{col_letter}{r}"]
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.fill = PatternFill("solid", fgColor=bg)

# RAG MFC
for col in ["F","I","M","P","U","X"]:
    rag_mfc(ws_ctrl, col, 4, 8)

# MFC date dépassée (rouge) et proche (orange) pour colonnes ⏱
red_fill   = PatternFill("solid", fgColor=RED_LIGHT)
orange_fill= PatternFill("solid", fgColor=ORANGE_LT)
for col in ["G","K","V","Z","AD"]:
    rng = f"{col}4:{col}8"
    ws_ctrl.conditional_formatting.add(rng,
        FormulaRule(formula=[f'AND(ISNUMBER({col}4),{col}4<TODAY())'], fill=red_fill))
    ws_ctrl.conditional_formatting.add(rng,
        FormulaRule(formula=[f'AND(ISNUMBER({col}4),{col}4>=TODAY(),{col}4<=TODAY()+30)'], fill=orange_fill))

# Number format for date columns
date_cols = ["D","G","H","K","S","V","W","Z","AA","AD"]
for col in date_cols:
    for r in range(4, 9):
        ws_ctrl[f"{col}{r}"].number_format = "DD/MM/YYYY"

# Dropdowns CONTRÔLES
for col in ["L","O"]:
    add_dv(ws_ctrl, '"Oui,Non,En attente"', f"{col}4:{col}200")
for col in ["R","T","W","AC"]:
    add_dv(ws_ctrl, '"Oui,Non,En cours"', f"{col}4:{col}200")

add_dv(ws_ctrl, '"Satisfaisant,Réserves mineures,Réserves majeures,Non réalisé"', "AB4:AB200")

thin_border(ws_ctrl, 2, 50, 1, 30)

# ────────────────────────────────────────────────────────────────────────────
# 5. PS (Prestations Simples)
# ────────────────────────────────────────────────────────────────────────────
ws_ps = wb.create_sheet("PS")
ws_ps.sheet_view.showGridLines = False

ws_ps.merge_cells("A1:L1")
hdr(ws_ps["A1"], "SUIVI PRESTATIONS SIMPLES (PS) — Registre allégé", bg=NAVY, fg=WHITE, sz=12)

ws_ps.merge_cells("A2:L2")
c = ws_ps["A2"]
c.value = ("ℹ  Les Prestations Simples font l'objet d'un suivi allégé : "
           "QECI obligatoire, DO requis pour nouveaux prestataires uniquement. "
           "Contrôles L1-06 et L1-07 applicables.")
c.font = Font(name="Arial", size=8, italic=True, color=NAVY)
c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_ps.row_dimensions[2].height = 28

ps_cols = [
    ("A","Prestataire",26), ("B","Fonction externalisée",28), ("C","Pays",12),
    ("D","Date contrat",14), ("E","Responsable interne",22),
    ("F","QECI présent ?",12), ("G","DO requis ?",12), ("H","DO présent ?",12),
    ("I","Contrat archivé EASY ?",16), ("J","Dernière vérif. EASY",14),
    ("K","Statut global",10), ("L","Commentaires",30),
]
for col, title, width in ps_cols:
    subhdr(ws_ps[f"{col}3"], title)
    ws_ps.column_dimensions[col].width = width
ws_ps.row_dimensions[3].height = 32

for r in range(4, 10):
    ws_ps.row_dimensions[r].height = 20
    bg = WHITE if r % 2 == 0 else GREY_BG
    for col in [c[0] for c in ps_cols]:
        cell = ws_ps[f"{col}{r}"]
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(name="Arial", size=9)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws_ps[f"K{r}"] = f'=IF(AND(F{r}="Oui",I{r}="Oui"),"Vert",IF(OR(F{r}="Non",I{r}="Non"),"Rouge","Orange"))'

add_dv(ws_ps, '"Oui,Non,En cours"', "F4:F200")
add_dv(ws_ps, '"Obligatoire (nouveau),N/A (existant)"', "G4:G200")
add_dv(ws_ps, '"Oui,Non,N/A"', "H4:H200")
add_dv(ws_ps, '"Oui,Non,En cours"', "I4:I200")
rag_mfc(ws_ps, "K", 4, 50)
thin_border(ws_ps, 2, 50, 1, 12)

# ────────────────────────────────────────────────────────────────────────────
# 6. SYNTHÈSE
# ────────────────────────────────────────────────────────────────────────────
ws_syn = wb.create_sheet("SYNTHÈSE")
ws_syn.sheet_view.showGridLines = False

ws_syn.merge_cells("A1:H1")
hdr(ws_syn["A1"], "SYNTHÈSE REMÉDIATION PCI — Tableau de bord",
    bg=NAVY, fg=WHITE, sz=13)

ws_syn.merge_cells("A2:H2")
ws_syn["A2"].value = f"Mise à jour : {datetime.date.today().strftime('%d/%m/%Y')}  |  Pilote : Opérations & Conformité Externalisations  |  Amundi Immobilier"
ws_syn["A2"].font = Font(name="Arial", size=9, italic=True, color=NAVY)
ws_syn["A2"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
ws_syn["A2"].alignment = Alignment(horizontal="left", vertical="center")
ws_syn.row_dimensions[2].height = 20

syn_hdrs = [
    ("A4","Prestataire"), ("B4","Type"), ("C4","Docs\n(RAG)"),
    ("D4","EASY\n(RAG)"), ("E4","Contrôles\nL1-08 S1"), ("F4","Contrôles\nL1-09 PCA"),
    ("G4","Risque global\n(estimé)"), ("H4","Prochaine\naction"),
]
for cell_ref, title in syn_hdrs:
    subhdr(ws_syn[cell_ref], title)
    col = cell_ref[0]
    ws_syn.column_dimensions[col].width = 24 if col in ["A","H"] else 12
ws_syn.row_dimensions[4].height = 34

for i in range(5):
    r = i + 5
    ws_syn.row_dimensions[r].height = 22
    bg = WHITE if r % 2 == 0 else GREY_BG
    ws_syn[f"A{r}"] = f"=DATABASE!A{i+4}"
    ws_syn[f"B{r}"] = f"=DATABASE!E{i+4}"
    ws_syn[f"C{r}"] = f'=IFERROR(DOCUMENTS!H{i+5},"—")'
    ws_syn[f"D{r}"] = f'=IFERROR(\'CONTRÔLES LOD1\'!F{i+4},"—")'
    ws_syn[f"E{r}"] = f'=IFERROR(\'CONTRÔLES LOD1\'!M{i+4},"—")'
    ws_syn[f"F{r}"] = f'=IFERROR(\'CONTRÔLES LOD1\'!U{i+4},"—")'
    ws_syn[f"G{r}"] = f'=IF(OR(C{r}="Rouge",D{r}="Rouge",E{r}="Rouge",F{r}="Rouge"),"Rouge",IF(OR(C{r}="Orange",D{r}="Orange",E{r}="Orange",F{r}="Orange"),"Orange","Vert"))'
    ws_syn[f"H{r}"] = ""
    for col in "ABCDEFGH":
        c = ws_syn[f"{col}{r}"]
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill("solid", fgColor=bg)

for col in ["C","D","E","F","G"]:
    rag_mfc(ws_syn, col, 5, 9)

thin_border(ws_syn, 4, 50, 1, 8)

# ────────────────────────────────────────────────────────────────────────────
# Onglet order & tab colors
# ────────────────────────────────────────────────────────────────────────────
tab_colors = {
    "DATABASE":       "001C4B",
    "DOCUMENTS":      "009EE0",
    "EASY":           "66B3DB",
    "CONTRÔLES LOD1": "001C4B",
    "PS":             "99CCE7",
    "SYNTHÈSE":       "FF9900",
}
for sheet_name, color in tab_colors.items():
    wb[sheet_name].sheet_properties.tabColor = color

# Freeze panes
for ws in [ws_db, ws_doc, ws_ctrl, ws_ps, ws_syn]:
    ws.freeze_panes = "A4"
ws_easy.freeze_panes = "D3"

output = "/home/claude/Matrice_PCI_Remediation.xlsx"
wb.save(output)
print(f"Saved: {output}")
