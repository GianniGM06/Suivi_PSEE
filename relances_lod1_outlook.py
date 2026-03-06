# =============================================================================
#  RELANCES AUTOMATIQUES LOD1 — L1-08 (Événements Majeurs) & L1-09 (Tests PCA)
#  Amundi Immobilier — Opérations & Conformité Externalisations
#
#  Prérequis  : Windows + Outlook installé + pip install openpyxl pywin32
#  Usage      : python relances_lod1_outlook.py
#               Le script lit la matrice Excel, détecte les non-répondants
#               et crée des brouillons Outlook (item.Display() — aucun envoi auto).
#
#  Calendrier de relance (défini dans la PROCÉDURE ci-dessous) :
#   • J-7  avant deadline  → Relance 1 (courtoises)
#   • J-3  avant deadline  → Relance 2 (urgente)
#   • J-1  avant deadline  → Relance 3 (dernière chance)
# =============================================================================

# ─── PROCÉDURE L1-08 & L1-09 ─────────────────────────────────────────────────
#
#  EXT-L1-08 | PILOTAGE DES ÉVÉNEMENTS MAJEURS (Semestriel)
#  ──────────────────────────────────────────────────────────
#  Périmètre  : Toutes les PCI Hors Groupe
#  Destinataires : Responsable opérationnel interne (champ L de DATABASE)
#
#  Calendrier S1 :
#    01 juin       → Lancement : envoi du mail initial à tous les responsables métier
#    23 juin       → Relance J-7  (deadline 30 juin)
#    27 juin       → Relance J-3
#    29 juin       → Relance J-1
#    30 juin       → DEADLINE — clôture S1
#
#  Calendrier S2 :
#    01 décembre   → Lancement : envoi du mail initial
#    23 décembre   → Relance J-7  (deadline 30 décembre)
#    27 décembre   → Relance J-3
#    29 décembre   → Relance J-1
#    30 décembre   → DEADLINE — clôture S2
#
#  EXT-L1-09 | REVUE TESTS PCA (Annuel — T1)
#  ──────────────────────────────────────────
#  Périmètre  : Toutes les PCI (Intra et Hors Groupe)
#  Destinataires : Contact PCA prestataire (champ N de DATABASE)
#
#  Calendrier :
#    15 janvier    → Lancement : demande du bilan PCA à chaque prestataire
#    24 mars       → Relance J-7  (deadline 31 mars)
#    28 mars       → Relance J-3
#    30 mars       → Relance J-1
#    31 mars       → DEADLINE — clôture annuelle T1
#
#  Le script détecte automatiquement à quelle étape de relance nous sommes
#  selon la date du jour et ne génère des brouillons que pour les
#  prestataires n'ayant pas encore rendu leur évaluation / bilan.
# ─────────────────────────────────────────────────────────────────────────────

import sys
import datetime
from pathlib import Path

try:
    import win32com.client as win32
except ImportError:
    sys.exit("❌ pywin32 non installé. Lancez : pip install pywin32")

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("❌ openpyxl non installé. Lancez : pip install openpyxl")

# ── Configuration ─────────────────────────────────────────────────────────────

MATRICE_PATH = Path(__file__).parent / "Matrice_PCI_Remediation.xlsx"

EXPEDITEUR_NOM   = "Gianni"                          # ← adapter
EXPEDITEUR_EMAIL = "gianni.xxx@amundi.com"           # ← adapter (affiché en signature)
DIRECTION        = "Opérations & Conformité Externalisations — Amundi Immobilier"

# Deadlines fixes (jour, mois) pour chaque campagne
DEADLINES = {
    "L108_S1": datetime.date(datetime.date.today().year, 6, 30),
    "L108_S2": datetime.date(datetime.date.today().year, 12, 30),
    "L109":    datetime.date(datetime.date.today().year, 3, 31),
}

# Seuils de relance en jours avant la deadline
SEUILS_RELANCE = [7, 3, 1]   # J-7, J-3, J-1

TODAY = datetime.date.today()


# ── Lecture de la matrice ──────────────────────────────────────────────────────

def lire_prestataires(path: Path) -> list[dict]:
    """Lit l'onglet DATABASE et retourne la liste des PCI actifs."""
    wb = load_workbook(path, data_only=True)
    ws = wb["DATABASE"]
    prestataires = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        nom, code, _, pays, type_p, intra_hg, crit, fonction, \
            date_debut, date_ech, resp_nom, resp_email, \
            contact_pca_nom, contact_pca_email, *_ = (list(row) + [None]*25)[:20]
        if not nom:
            break
        if str(type_p).strip().upper() != "PCI":
            continue
        prestataires.append({
            "nom":               str(nom).strip(),
            "intra_hg":          str(intra_hg).strip() if intra_hg else "",
            "resp_nom":          str(resp_nom).strip() if resp_nom else "",
            "resp_email":        str(resp_email).strip() if resp_email else "",
            "contact_pca_nom":   str(contact_pca_nom).strip() if contact_pca_nom else "",
            "contact_pca_email": str(contact_pca_email).strip() if contact_pca_email else "",
        })
    wb.close()
    return prestataires


def lire_statuts_controles(path: Path) -> dict:
    """
    Lit l'onglet CONTRÔLES LOD1 et retourne un dict :
      { nom_prestataire: { "L108_S1": bool_rendu, "L108_S2": bool_rendu, "L109": bool_rendu } }
    """
    wb = load_workbook(path, data_only=True)
    ws = wb["CONTRÔLES LOD1"]
    statuts = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0]:
            break
        nom   = str(row[0]).strip()
        # Colonnes : L=11 (S1 reporting reçu), O=14 (S2), R=17 (L109 bilan reçu)
        # openpyxl index 0-based → col L = index 11, O = 14, R = 17
        s1_recu  = str(row[11]).strip().lower() if row[11] else ""
        s2_recu  = str(row[14]).strip().lower() if row[14] else ""
        pca_recu = str(row[17]).strip().lower() if row[17] else ""
        statuts[nom] = {
            "L108_S1": s1_recu == "oui",
            "L108_S2": s2_recu == "oui",
            "L109":    pca_recu == "oui",
        }
    wb.close()
    return statuts


# ── Logique de relance ─────────────────────────────────────────────────────────

def detecter_relances_actives() -> list[tuple[str, int]]:
    """
    Retourne la liste des (campagne, jours_avant_deadline) actives aujourd'hui.
    Ex : [("L108_S1", 7), ("L109", 3)]
    """
    actives = []
    for campagne, deadline in DEADLINES.items():
        delta = (deadline - TODAY).days
        if delta in SEUILS_RELANCE:
            actives.append((campagne, delta))
    return actives


def libelle_campagne(campagne: str) -> str:
    return {
        "L108_S1": "EXT-L1-08 — Pilotage Événements Majeurs — S1",
        "L108_S2": "EXT-L1-08 — Pilotage Événements Majeurs — S2",
        "L109":    "EXT-L1-09 — Revue Tests PCA",
    }.get(campagne, campagne)


def semestre_label(campagne: str) -> str:
    return {
        "L108_S1": "Semestre 1 (clôture 30 juin)",
        "L108_S2": "Semestre 2 (clôture 30 décembre)",
        "L109":    "Campagne annuelle T1 (clôture 31 mars)",
    }.get(campagne, "")


def rang_relance(jours: int) -> str:
    return {7: "1ère relance (J-7)", 3: "2ème relance (J-3)", 1: "Dernière relance (J-1)"}.get(jours, "Relance")


# ── Corps des mails ────────────────────────────────────────────────────────────

def corps_l108(prestataire: dict, campagne: str, jours: int) -> tuple[str, str]:
    """Retourne (sujet, corps HTML) pour L1-08."""
    deadline = DEADLINES[campagne]
    sujet = (
        f"[{rang_relance(jours)}] {libelle_campagne(campagne)} — "
        f"{prestataire['nom']} — Transmission attendue avant le {deadline.strftime('%d/%m/%Y')}"
    )
    corps = f"""
<p>Bonjour {prestataire['resp_nom']},</p>

<p>Dans le cadre du dispositif de contrôle de premier niveau des externalisations
(EXT-L1-08 — Pilotage des événements majeurs), nous vous rappelons que votre
évaluation semestrielle concernant la prestation <strong>{prestataire['nom']}</strong>
est attendue pour le <strong>{deadline.strftime('%d %B %Y')}</strong>
({semestre_label(campagne)}).</p>

<p>À ce jour, votre évaluation n'a pas encore été transmise dans la matrice de suivi.</p>

<p><strong>Ce qui vous est demandé :</strong></p>
<ul>
  <li>Confirmer la réception du reporting du prestataire sur la période (indicateurs qualité, incidents, SLA)</li>
  <li>Identifier et documenter tout événement majeur ayant conduit à une dégradation du service</li>
  <li>Renseigner le statut (Vert / Orange / Rouge) et un commentaire dans l'onglet <em>CONTRÔLES LOD1</em> de la matrice partagée</li>
</ul>

<p>Pour rappel, ce contrôle est obligatoire pour toutes les PCI Hors Groupe et
constitue une pièce justificative exigée dans le cadre de nos obligations
EBA/DORA et du Comité de Contrôle Interne (CCI).</p>

<p>En cas de difficulté ou d'absence d'événement majeur à signaler, merci de
le confirmer explicitement en renseignant <em>"Oui / Vert / RAS"</em> dans la matrice.</p>

<p>Je reste disponible pour toute question.</p>

<p>Bien cordialement,</p>
<p><strong>{EXPEDITEUR_NOM}</strong><br>
{DIRECTION}<br>
<em>{EXPEDITEUR_EMAIL}</em></p>
"""
    return sujet, corps


def corps_l109(prestataire: dict, jours: int) -> tuple[str, str]:
    """Retourne (sujet, corps HTML) pour L1-09 — destinataire = contact PCA prestataire."""
    deadline = DEADLINES["L109"]
    sujet = (
        f"[{rang_relance(jours)}] {libelle_campagne('L109')} — "
        f"Transmission du bilan PCA {deadline.year} attendue avant le {deadline.strftime('%d/%m/%Y')}"
    )
    corps = f"""
<p>Bonjour {prestataire['contact_pca_nom']},</p>

<p>Dans le cadre de nos obligations contractuelles et réglementaires
(EBA Guidelines, DORA — EXT-L1-09), nous sollicitons la transmission
du <strong>bilan de tests PCA</strong> relatif à votre prestation pour
Amundi Immobilier, pour l'exercice <strong>{deadline.year - 1}</strong>.</p>

<p>Ce document est attendu pour le <strong>{deadline.strftime('%d %B %Y')}</strong>
au plus tard.</p>

<p><strong>Documents attendus :</strong></p>
<ul>
  <li>Bilan des tests PCA réalisés en {deadline.year - 1} (scénarios, résultats, anomalies éventuelles)</li>
  <li>Preuve de la réalisation des tests (rapport, PV de test ou attestation)</li>
  <li>Plan d'action correctif si des anomalies ont été identifiées</li>
  <li>Confirmation de l'interaction entre votre dispositif de gestion de crise et celui d'Amundi Immobilier</li>
</ul>

<p>Ces éléments constituent la piste d'audit requise pour notre contrôle interne
et sont susceptibles d'être demandés par nos autorités de supervision (BCE, ACPR/AMF).</p>

<p>Merci de transmettre ces documents à l'adresse suivante :
<strong>{EXPEDITEUR_EMAIL}</strong> en indiquant en objet la référence de votre contrat.</p>

<p>En l'absence de retour avant la date limite, nous serons contraints d'enregistrer
ce contrôle en statut <em>Rouge</em> et d'en informer notre Direction des Risques.</p>

<p>Bien cordialement,</p>
<p><strong>{EXPEDITEUR_NOM}</strong><br>
{DIRECTION}<br>
<em>{EXPEDITEUR_EMAIL}</em></p>
"""
    return sujet, corps


# ── Création des brouillons Outlook ───────────────────────────────────────────

def creer_brouillon(outlook, destinataire_email: str, sujet: str, corps_html: str):
    """Crée un brouillon Outlook et l'affiche via .Display()."""
    mail = outlook.CreateItem(0)  # 0 = olMailItem
    mail.To      = destinataire_email
    mail.Subject = sujet
    mail.HTMLBody = corps_html
    mail.Display(False)  # False = non modal (n'attend pas fermeture)
    return mail


# ── Point d'entrée ─────────────────────────────────────────────────────────────

def main():
    print(f"\n{'='*70}")
    print(f"  RELANCES LOD1 — Amundi Immobilier  |  {TODAY.strftime('%d/%m/%Y')}")
    print(f"{'='*70}\n")

    # 1. Vérifier le fichier matrice
    if not MATRICE_PATH.exists():
        sys.exit(f"❌ Fichier introuvable : {MATRICE_PATH}\n   Vérifiez le chemin dans la variable MATRICE_PATH.")

    # 2. Lire les données
    print("📂 Lecture de la matrice Excel...")
    prestataires = lire_prestataires(MATRICE_PATH)
    statuts      = lire_statuts_controles(MATRICE_PATH)
    print(f"   → {len(prestataires)} prestataires PCI chargés\n")

    # 3. Détecter les campagnes actives aujourd'hui
    relances_actives = detecter_relances_actives()

    if not relances_actives:
        print("✅ Aucune relance à envoyer aujourd'hui.")
        print(f"\n   Prochaines deadlines :")
        for camp, dl in DEADLINES.items():
            delta = (dl - TODAY).days
            print(f"   • {libelle_campagne(camp)} → {dl.strftime('%d/%m/%Y')} (dans {delta} jours)")
        return

    # 4. Ouvrir Outlook
    print("📧 Connexion à Outlook...")
    try:
        outlook = win32.Dispatch("Outlook.Application")
    except Exception as e:
        sys.exit(f"❌ Impossible d'ouvrir Outlook : {e}")

    total_brouillons = 0

    for campagne, jours in relances_actives:
        print(f"\n{'─'*60}")
        print(f"  {libelle_campagne(campagne)} — {rang_relance(jours)}")
        print(f"  Deadline : {DEADLINES[campagne].strftime('%d/%m/%Y')}")
        print(f"{'─'*60}")

        for p in prestataires:
            nom = p["nom"]
            statut_prest = statuts.get(nom, {})

            if campagne in ["L108_S1", "L108_S2"]:
                # L1-08 : destinataire = responsable opérationnel interne
                if statut_prest.get(campagne, False):
                    print(f"   ✅ {nom} — déjà rendu, pas de relance")
                    continue
                if not p["resp_email"] or "@" not in p["resp_email"]:
                    print(f"   ⚠️  {nom} — email responsable manquant, relance ignorée")
                    continue
                sujet, corps = corps_l108(p, campagne, jours)
                creer_brouillon(outlook, p["resp_email"], sujet, corps)
                print(f"   📬 Brouillon créé → {p['resp_nom']} ({p['resp_email']})")
                total_brouillons += 1

            elif campagne == "L109":
                # L1-09 : destinataire = contact PCA prestataire
                if statut_prest.get("L109", False):
                    print(f"   ✅ {nom} — bilan PCA déjà reçu, pas de relance")
                    continue
                if not p["contact_pca_email"] or "@" not in p["contact_pca_email"]:
                    print(f"   ⚠️  {nom} — email contact PCA manquant, relance ignorée")
                    continue
                sujet, corps = corps_l109(p, jours)
                creer_brouillon(outlook, p["contact_pca_email"], sujet, corps)
                print(f"   📬 Brouillon créé → {p['contact_pca_nom']} ({p['contact_pca_email']})")
                total_brouillons += 1

    print(f"\n{'='*70}")
    print(f"  ✅ {total_brouillons} brouillon(s) Outlook créé(s) — vérifiez avant envoi.")
    print(f"{'='*70}\n")


if __name__ == "__main__":
    main()
